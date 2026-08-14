import json
import unittest

import pandas as pd
from openpyxl import load_workbook

import delivery_audit_backfill
import delivery_match_adapter
import delivery_reference
import delivery_runtime
import delivery_stage1_adapter
import delivery_workflow
import processors
import tool_common


class MultiUnloadAverageTests(unittest.TestCase):
    def setUp(self):
        self.rows = pd.DataFrame(
            [
                {
                    "仓库": "LA", "统计周期": "2026-W28", "专线线路": "LA-NJ",
                    "是否FTL发车": True, "出库体积": 50, "出库卡板数": 2,
                    "派送成本": 500, "派送时效": 2, "匹配备注集合": "正常",
                    "批次号集合": "A", "车次号": "T1", "出库类型": "调拨",
                    "调入仓库": "NJ", "业务场景": "仓间调拨",
                },
                {
                    "仓库": "LA", "统计周期": "2026-W28", "专线线路": "LA-NJ",
                    "是否FTL发车": True, "出库体积": 60, "出库卡板数": 4,
                    "派送成本": 600, "派送时效": 20, "匹配备注集合": "里外两卸",
                    "批次号集合": "B", "车次号": "T2", "出库类型": "调拨",
                    "调入仓库": "NJ", "业务场景": "仓间调拨",
                },
            ]
        )

    def test_slc3_fba_reference_uses_confirmed_shipping_address(self):
        reference_df, reference_map = delivery_reference.load_fba_reference()

        self.assertIn("SLC3", reference_map)
        slc3 = reference_map["SLC3"]
        self.assertEqual(slc3["目的地示例"], "Amazon-SLC3")
        self.assertEqual(
            slc3["地址"],
            "355 N John Glenn Rd, Salt Lake City, UT 84116-4413",
        )
        self.assertEqual(slc3["邮编"], "84116")
        self.assertEqual(slc3["邮编前三位"], "841")
        self.assertEqual(slc3["州"], "UT")
        self.assertEqual(slc3["站点名称"], "Amazon/SLC3")

        matched = delivery_reference.match_fba_reference("Amazon SLC3")
        self.assertEqual(matched["代码"], "SLC3")
        self.assertEqual(matched["邮编"], "84116")
        self.assertEqual(matched["州"], "UT")

    def test_linehaul_totals_include_both_but_averages_exclude_marked_row(self):
        report = delivery_audit_backfill._build_linehaul_sheet(self.rows)
        row = report.loc[report["专线线路"] == "LA-NJ"].iloc[0]
        self.assertEqual(row["车次数"], 2)
        self.assertEqual(row["总出库体积"], 110)
        self.assertEqual(row["总派送成本"], 1100)
        self.assertEqual(row["平均整车价"], 500)
        self.assertEqual(row["每方平均价"], 10)
        self.assertEqual(row["平均每车出库体积"], 50)
        self.assertEqual(row["平均派送时效"], 2)
        self.assertEqual(row["P80派送时效"], 2)

    def test_transfer_totals_include_both_but_averages_exclude_marked_row(self):
        report = delivery_runtime._build_transfer_report(self.rows)
        row = report.iloc[0]
        self.assertEqual(row["车次数"], 2)
        self.assertEqual(row["总出库体积"], 110)
        self.assertEqual(row["总派送成本"], 1100)
        self.assertEqual(row["平均整车价"], 500)
        self.assertEqual(row["每方平均价"], 10)
        self.assertEqual(row["平均每车出库体积"], 50)

    def test_blank_transfer_fields_do_not_turn_ordinary_linehaul_into_transfer(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-NJ",
                "是否FTL发车": True, "车次号": "T-NORMAL", "批次号集合": "NORMAL",
                "批次车份额": 1, "出库体积": 60, "出库卡板数": 10, "派送成本": 9500,
                "出库类型": "派送", "业务场景": "普通派送", "调入仓库": float("nan"),
                "调拨目标仓代码": float("nan"), "调拨覆盖审核": float("nan"),
                "系统产品类型": "FBA", "主产品类型": "FBA", "批次目的地类型": "FBA",
                "邮编来源": "内置FBA仓点邮编表",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-NJ",
                "是否FTL发车": True, "车次号": "T-TRANSFER", "批次号集合": "TRANSFER",
                "批次车份额": 1, "出库体积": 80, "出库卡板数": 20, "派送成本": 9700,
                "出库类型": "调拨", "业务场景": "仓间调拨", "调入仓库": "新泽西盈仓",
                "调拨目标仓代码": "NJ", "调拨覆盖审核": "同批次调拨统一覆盖:新泽西盈仓",
                "系统产品类型": "仓间调拨", "主产品类型": "仓间调拨", "批次目的地类型": "其他",
                "邮编来源": "仓间调拨目标仓地址",
            },
        ])

        report = delivery_runtime._build_transfer_report(rows)

        self.assertEqual(len(report), 1)
        self.assertEqual(report.iloc[0]["调拨目标仓"], "NJ盈仓")
        self.assertEqual(report.iloc[0]["车次数"], 1)
        self.assertEqual(report.iloc[0]["总出库体积"], 80)
        self.assertEqual(report.iloc[0]["总派送成本"], 9700)

    def test_either_marker_is_sufficient(self):
        for marker in ("里", "外", "里外"):
            rows = self.rows.copy()
            rows.loc[1, "匹配备注集合"] = marker
            report = delivery_audit_backfill._build_linehaul_sheet(rows)
            row = report.loc[report["专线线路"] == "LA-NJ"].iloc[0]
            self.assertEqual(row["平均派送时效"], 2)

    def test_stage1_and_stage2_keep_batch_remarks_as_last_column(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1", "批次号": "A",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03", "出库体积": 10, "出库卡板数": 2,
                "派送成本": 100, "FBA/FBX": "FBX", "平台名称": "谷仓", "FBX代码": "16号仓", "备注": "正常批次",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1", "批次号": "B",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03", "出库体积": 30, "出库卡板数": 4,
                "派送成本": 600, "FBA/FBX": "FBX", "平台名称": "谷仓", "FBX代码": "16号仓", "备注": "里仓两卸",
            },
        ])
        stage1 = delivery_workflow.build_cleaned_batches_from_detail(detail)
        self.assertEqual(len(stage1), 2)
        self.assertEqual(stage1.columns[-1], "备注")
        self.assertEqual(set(stage1["备注"]), {"正常批次", "里仓两卸"})

        delivery_runtime.bootstrap(delivery_workflow)
        stage2 = delivery_workflow.prepare_stage2_for_report(stage1, pd.DataFrame(), "按周统计")
        self.assertEqual(stage2.columns[-1], "同车次备注集合")
        self.assertEqual(set(stage2["同车次备注集合"]), {"正常批次", "里仓两卸"})

        exported = delivery_match_adapter._finalize_sheet(stage2, "明细")
        self.assertEqual(exported.columns[-1], "同车次备注集合")
        self.assertEqual(set(exported["同车次备注集合"]), {"正常批次", "里仓两卸"})

    def test_creator_is_preserved_as_audit_field_through_stage1_and_stage2(self):
        raw = pd.DataFrame([
            {
                "仓库": "LA", "派送方式": "卡车派送", "运输类型": "FTL",
                "车次号": "T-CREATOR", "批次号": "B-CREATOR",
                "创建时间": "2026-07-01", "Creator": "Alice",
                "出库时间": "2026-07-02", "签收时间": "2026-07-03",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 200,
                "目的地": "Amazon-ONT8", "车型": "53尺大车", "装车类型": "卡板",
            },
            {
                "仓库": "LA", "派送方式": "卡车派送", "运输类型": "FTL",
                "车次号": "T-CREATOR", "批次号": "B-CREATOR",
                "创建时间": "2026-07-01", "Creator": "Bob",
                "出库时间": "2026-07-02", "签收时间": "2026-07-03",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 200,
                "目的地": "Amazon-ONT8", "车型": "53尺大车", "装车类型": "卡板",
            },
        ])

        detail, _, _ = processors.process_delivery_stage1_from_df(raw, "LA")
        self.assertIn("创建人", detail.columns)
        self.assertEqual(detail["创建人"].tolist(), ["Alice", "Bob"])

        detail = delivery_stage1_adapter.repair_delivery_stage1_numeric_columns(detail)
        stage1 = delivery_workflow.build_cleaned_batches_from_detail(detail)
        self.assertEqual(stage1.iloc[0]["创建人"], "Alice,Bob")

        delivery_runtime.bootstrap(delivery_workflow)
        stage2 = delivery_workflow.prepare_stage2_for_report(stage1, pd.DataFrame(), "按月统计")
        self.assertEqual(stage2.iloc[0]["创建人"], "Alice,Bob")

        exported = delivery_match_adapter._finalize_sheet(stage2, "明细")
        self.assertIn("创建人", exported.columns)
        self.assertEqual(exported.iloc[0]["创建人"], "Alice,Bob")
        self.assertEqual(exported.columns[-1], "同车次备注集合")

    def test_creator_column_is_kept_in_detail_export_when_source_values_are_blank(self):
        detail = pd.DataFrame([{
            "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL",
            "车次号": "T-BLANK-CREATOR", "批次号": "B-BLANK-CREATOR", "创建人": "",
            "出库时间": "2026-07-01", "签收时间": "2026-07-02",
            "出库体积": 40, "出库卡板数": 10, "派送成本": 300,
            "FBA/FBX": "FBA", "FBA仓点代码": "ONT8", "备注": "",
        }])

        stage1 = delivery_workflow.build_cleaned_batches_from_detail(detail)
        delivery_runtime.bootstrap(delivery_workflow)
        stage2 = delivery_workflow.prepare_stage2_for_report(stage1, pd.DataFrame(), "按月统计")
        exported = delivery_match_adapter._finalize_sheet(stage2, "明细")

        self.assertIn("创建人", exported.columns)
        self.assertEqual(exported.iloc[0]["创建人"], "")

    def test_transfer_and_fba_batches_in_same_trip_keep_individual_destinations(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "TRANSFER-TRIP-1", "批次号": "A",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03",
                "出库体积": 20, "出库卡板数": 2, "派送成本": 300,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "目的地": "Amazon-ONT8", "出库类型": "调拨",
                "业务场景": "仓间调拨", "调入仓库": "NJ", "备注": "",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "TRANSFER-TRIP-1", "批次号": "B",
                "出库时间": "2026-07-01", "签收时间": "2026-07-04",
                "出库体积": 40, "出库卡板数": 4, "派送成本": 600,
                "FBA/FBX": "FBA", "FBA仓点代码": "LAS1",
                "目的地": "Amazon-LAS1", "出库类型": "派送",
                "业务场景": "", "调入仓库": "", "备注": "",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)

        self.assertEqual(len(cleaned), 2)
        transfer_batch = cleaned.loc[cleaned["批次号"] == "A"].iloc[0]
        fba_batch = cleaned.loc[cleaned["批次号"] == "B"].iloc[0]
        self.assertEqual(transfer_batch["出库类型"], "调拨")
        self.assertEqual(transfer_batch["调入仓库"], "新泽西盈仓")
        self.assertEqual(transfer_batch["调拨目标仓代码"], "NJ")
        self.assertEqual(transfer_batch["主产品类型"], "仓间调拨")
        self.assertEqual(transfer_batch["批次目的地类型"], "其他")
        self.assertEqual(transfer_batch["批次目的仓点"], "新泽西盈仓")
        self.assertEqual(transfer_batch["FBA出库体积"], 0)
        self.assertTrue(transfer_batch["调拨覆盖审核"].startswith("同批次调拨统一覆盖"))
        self.assertEqual(fba_batch["出库类型"], "派送")
        self.assertEqual(fba_batch["主产品类型"], "FBA")
        self.assertEqual(fba_batch["批次目的地类型"], "FBA")
        self.assertEqual(fba_batch["批次目的仓点"], "LAS1")
        self.assertEqual(fba_batch["FBA仓点代码集合"], "LAS1")
        self.assertEqual(fba_batch["FBA出库体积"], 40)
        self.assertNotIn("新泽西盈仓", fba_batch["目的仓点分配明细"])

        delivery_runtime.bootstrap(delivery_workflow)
        matched = delivery_workflow.prepare_stage2_for_report(
            cleaned,
            pd.DataFrame(),
            "按月统计",
        )
        self.assertEqual(set(matched["主产品类型"]), {"仓间调拨", "FBA"})
        self.assertEqual(matched.loc[matched["批次号"] == "A", "专线线路"].iloc[0], "LA-NJ")
        transfer = delivery_runtime._build_transfer_report(matched)
        self.assertEqual(len(transfer), 1)
        self.assertEqual(transfer.iloc[0]["总出库体积"], 20)
        self.assertNotIn("批次号集合", transfer.columns)
        self.assertNotIn("车次号集合", transfer.columns)
        fba_rank = delivery_match_adapter.build_fba_rank_sheet(matched)
        self.assertEqual(fba_rank.iloc[0]["FBA仓点"], "LAS1")
        self.assertEqual(fba_rank.iloc[0]["出库体积"], 40)

    def test_transfer_without_trip_propagates_within_same_batch(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "", "批次号": "BATCH-SAV",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03",
                "出库体积": 10, "出库卡板数": 1, "派送成本": 100,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "出库类型": "调拨", "调入仓库": "SAV", "备注": "",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "", "批次号": "BATCH-SAV",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03",
                "出库体积": 5, "出库卡板数": 1, "派送成本": 100,
                "FBA/FBX": "FBA", "FBA仓点代码": "LAS1",
                "出库类型": "派送", "调入仓库": "", "备注": "",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)

        self.assertEqual(len(cleaned), 1)
        self.assertEqual(cleaned.iloc[0]["调入仓库"], "萨凡纳盈仓")
        self.assertEqual(cleaned.iloc[0]["批次目的仓点"], "萨凡纳盈仓")
        self.assertTrue(cleaned.iloc[0]["调拨覆盖审核"].startswith("同批次调拨统一覆盖"))

    def test_conflicting_or_missing_transfer_target_is_invalid(self):
        base = {
            "仓库": "LA", "标准运输类型": "FTL", "车次号": "CONFLICT-1",
            "出库时间": "2026-07-01", "签收时间": "2026-07-03",
            "出库体积": 20, "出库卡板数": 2, "派送成本": 300,
            "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
            "出库类型": "调拨", "业务场景": "仓间调拨", "备注": "",
        }
        separate_batches = pd.DataFrame([
            {**base, "原始行号": 2, "批次号": "A", "调入仓库": "NJ"},
            {**base, "原始行号": 3, "批次号": "B", "调入仓库": "DAL"},
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(separate_batches)
        self.assertEqual(len(cleaned), 2)
        self.assertEqual(set(cleaned["批次目的仓点"]), {"新泽西盈仓", "达拉斯盈仓"})

        same_batch_conflict = pd.DataFrame([
            {**base, "原始行号": 4, "批次号": "C", "调入仓库": "NJ"},
            {**base, "原始行号": 5, "批次号": "C", "调入仓库": "DAL"},
        ])
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(same_batch_conflict)
        invalid = pd.DataFrame(cleaned.attrs["batch_invalid_records"])

        self.assertTrue(cleaned.empty)
        self.assertEqual(len(invalid), 1)
        self.assertIn("同批次调拨目标冲突", invalid.iloc[0]["批次无效原因"])

        missing = pd.DataFrame([{
            **base,
            "原始行号": 6,
            "车次号": "MISSING-TARGET",
            "批次号": "C",
            "调入仓库": "",
            "目的地": "Amazon-LAS1",
        }])
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(missing)
        invalid = pd.DataFrame(cleaned.attrs["batch_invalid_records"])
        self.assertTrue(cleaned.empty)
        self.assertIn("调拨目标盈仓无法识别", invalid.iloc[0]["批次无效原因"])

    def test_stage2_legacy_rows_reapply_batch_transfer_override(self):
        legacy = pd.DataFrame([
            {
                "仓库": "LA", "标准运输类型": "FTL", "车次号": "OLD-TRIP",
                "是否有真实车次号": True, "批次号": "A", "批次号集合": "A",
                "批次出库时间": "2026-07-01", "批次签收时间": "2026-07-03",
                "出库体积": 20, "出库卡板数": 2, "派送成本": 300,
                "批次车份额": 0.33, "FBA出库体积": 20, "FBX出库体积": 0,
                "系统产品类型": "FBA", "主产品类型": "FBA",
                "FBA仓点代码集合": "ONT8", "标准邮编集合": "92316",
                "出库类型": "调拨", "业务场景": "仓间调拨", "调入仓库": "DAL",
                "备注": "",
            },
            {
                "仓库": "LA", "标准运输类型": "FTL", "车次号": "OLD-TRIP",
                "是否有真实车次号": True, "批次号": "B", "批次号集合": "B",
                "批次出库时间": "2026-07-01", "批次签收时间": "2026-07-04",
                "出库体积": 40, "出库卡板数": 4, "派送成本": 600,
                "批次车份额": 0.67, "FBA出库体积": 40, "FBX出库体积": 0,
                "系统产品类型": "FBA", "主产品类型": "FBA",
                "FBA仓点代码集合": "LAS1", "标准邮编集合": "92316",
                "出库类型": "派送", "业务场景": "", "调入仓库": "",
                "备注": "",
            },
        ])

        delivery_runtime.bootstrap(delivery_workflow)
        matched = delivery_workflow.prepare_stage2_for_report(
            legacy,
            pd.DataFrame(),
            "按月统计",
        )

        transfer_batch = matched.loc[matched["批次号"] == "A"].iloc[0]
        fba_batch = matched.loc[matched["批次号"] == "B"].iloc[0]
        self.assertEqual(transfer_batch["调入仓库"], "达拉斯盈仓")
        self.assertEqual(transfer_batch["主产品类型"], "仓间调拨")
        self.assertEqual(transfer_batch["FBA仓点代码集合"], "")
        self.assertEqual(transfer_batch["专线线路"], "LA-DAL")
        self.assertEqual(fba_batch["调入仓库"], "")
        self.assertEqual(fba_batch["主产品类型"], "FBA")
        self.assertEqual(fba_batch["FBA仓点代码集合"], "LAS1")

    def test_ftl_big_truck_floor_loading_fee_uses_exact_batch_share_once(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "FLOOR-TRIP-1", "批次号": "A",
                "出库时间": "2026-07-01", "签收时间": "2026-07-03",
                "出库体积": 40, "出库卡板数": 8, "派送成本": 350,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "车型": "53尺大车", "装车类型": "地板", "备注": "",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL",
                "车次号": "FLOOR-TRIP-1", "批次号": "B",
                "出库时间": "2026-07-01", "签收时间": "2026-07-04",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 220,
                "FBA/FBX": "FBA", "FBA仓点代码": "LAS1",
                "车型": "53尺大车", "装车类型": "地板", "备注": "",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        batch_a = cleaned.loc[cleaned["批次号"] == "A"].iloc[0]
        batch_b = cleaned.loc[cleaned["批次号"] == "B"].iloc[0]
        self.assertAlmostEqual(batch_a["批次车份额"], 2 / 3)
        self.assertAlmostEqual(batch_b["批次车份额"], 1 / 3)
        self.assertEqual(batch_a["原始派送成本"], 350)
        self.assertEqual(batch_b["原始派送成本"], 220)
        self.assertAlmostEqual(batch_a["大车地板装车费"], 200 * 2 / 3)
        self.assertAlmostEqual(batch_b["大车地板装车费"], 200 * 1 / 3)
        self.assertAlmostEqual(batch_a["派送成本"], 350 + 200 * 2 / 3)
        self.assertAlmostEqual(batch_b["派送成本"], 220 + 200 * 1 / 3)
        self.assertAlmostEqual(cleaned["大车地板装车费"].sum(), 200)
        self.assertAlmostEqual(cleaned["派送成本"].sum(), 770)
        allocation = json.loads(batch_a["目的仓点分配明细"])[0]
        self.assertAlmostEqual(allocation["原始派送成本"], 350)
        self.assertAlmostEqual(allocation["大车地板装车费"], 200 * 2 / 3)
        self.assertAlmostEqual(allocation["派送成本"], 350 + 200 * 2 / 3)

        first = delivery_workflow.prepare_stage2_for_report(
            cleaned,
            pd.DataFrame([{"工作单号": "", "ZIP": ""}]),
            "按月统计",
        )
        second = delivery_workflow.prepare_stage2_for_report(
            first,
            pd.DataFrame([{"工作单号": "", "ZIP": ""}]),
            "按月统计",
        )
        self.assertAlmostEqual(first["派送成本"].sum(), 770)
        self.assertAlmostEqual(second["派送成本"].sum(), 770)
        self.assertAlmostEqual(second["大车地板装车费"].sum(), 200)

    def test_floor_loading_fee_excludes_pallet_small_truck_ltl_and_missing_share(self):
        rows = pd.DataFrame([
            {"标准运输类型": "FTL", "车型标准值": "53尺大车", "装车类型标准值": "卡板", "批次车份额": 1, "派送成本": 100},
            {"标准运输类型": "FTL", "车型标准值": "小车", "装车类型标准值": "地板", "批次车份额": 1, "派送成本": 200},
            {"标准运输类型": "LTL", "车型标准值": "53尺大车", "装车类型标准值": "地板", "批次车份额": 1, "派送成本": 300},
            {"标准运输类型": "FTL", "车型标准值": "53尺大车", "装车类型标准值": "地板", "批次车份额": pd.NA, "派送成本": 400},
        ])
        adjusted = tool_common.apply_floor_loading_fee(rows)
        self.assertEqual(adjusted["大车地板装车费"].sum(), 0)
        self.assertEqual(adjusted["派送成本"].tolist(), [100, 200, 300, 400])

        half_truck = tool_common.apply_floor_loading_fee(pd.DataFrame([{
            "标准运输类型": "FTL", "车型标准值": "53尺大车",
            "装车类型标准值": "地板", "批次车份额": 0.5, "派送成本": 400,
        }])).iloc[0]
        self.assertEqual(half_truck["大车地板装车费"], 100)
        self.assertEqual(half_truck["派送成本"], 500)

    def test_stage1_runtime_and_stage2_cost_reports_keep_floor_loading_fee(self):
        source = pd.DataFrame([
            {
                "装车类型": "地板", "批次号": "FLOOR-A", "出库体积": 40,
                "目的地": "Amazon-ONT8", "派送卡车": "JTeam INC",
                "车次号": "FLOOR-RUNTIME-1", "出库时间": "2026-07-01",
                "签收时间": "2026-07-03", "出库卡板数": 8,
                "派送方式": "卡车派送", "出库类型": "派送",
                "派送成本": 350, "运输类型": "FTL", "车型": "53尺大车",
            },
            {
                "装车类型": "地板", "批次号": "FLOOR-B", "出库体积": 20,
                "目的地": "Amazon-LAS1", "派送卡车": "JTeam INC",
                "车次号": "FLOOR-RUNTIME-1", "出库时间": "2026-07-01",
                "签收时间": "2026-07-04", "出库卡板数": 4,
                "派送方式": "卡车派送", "出库类型": "派送",
                "派送成本": 220, "运输类型": "FTL", "车型": "53尺大车",
            },
        ])

        delivery_runtime.bootstrap(delivery_workflow)
        cleaned, invalid, _, _ = delivery_workflow.process_stage1_raw_files_to_cleaned_batches(
            [("地板装车费样例.xlsx", source)],
            "LA",
        )
        self.assertTrue(invalid.empty)
        self.assertAlmostEqual(cleaned["原始派送成本"].sum(), 570)
        self.assertAlmostEqual(cleaned["大车地板装车费"].sum(), 200)
        self.assertAlmostEqual(cleaned["派送成本"].sum(), 770)

        reports = delivery_workflow.process_stage2_analysis(
            cleaned,
            pd.DataFrame([{"工作单号": "", "ZIP": ""}]),
            "按月统计",
        )
        self.assertAlmostEqual(reports["每方价格参考"]["总派送成本"].sum(), 770)
        self.assertAlmostEqual(
            reports["分类型价格参考"]["总派送成本"].sum(),
            770,
        )

    def test_stage1_remark_marker_excludes_entire_trip_from_linehaul_averages(self):
        rows = self.rows.copy()
        rows["同车次备注集合"] = ["正常", "外仓两卸"]
        rows["匹配备注集合"] = ""
        report = delivery_audit_backfill._build_linehaul_sheet(rows)
        row = report.loc[report["专线线路"] == "LA-NJ"].iloc[0]
        self.assertEqual(row["车次数"], 2)
        self.assertEqual(row["总出库体积"], 110)
        self.assertEqual(row["平均每车出库体积"], 50)
        self.assertEqual(row["平均派送时效"], 2)
        self.assertEqual(row["P80派送时效"], 2)

    def test_under_45_cbm_trip_is_totals_only_and_45_is_eligible(self):
        rows = self.rows.copy()
        rows["匹配备注集合"] = "正常"
        rows["同车次备注集合"] = "正常"
        rows[["出库体积", "派送成本", "派送时效"]] = rows[["出库体积", "派送成本", "派送时效"]].astype(float)
        rows.loc[0, ["出库体积", "派送成本", "派送时效"]] = [45, 450, 2]
        rows.loc[1, ["出库体积", "派送成本", "派送时效"]] = [44.99, 899.8, 20]

        linehaul = delivery_audit_backfill._build_linehaul_sheet(rows)
        linehaul_row = linehaul.loc[linehaul["专线线路"] == "LA-NJ"].iloc[0]
        self.assertEqual(linehaul_row["车次数"], 2)
        self.assertEqual(linehaul_row["总出库体积"], 89.99)
        self.assertEqual(linehaul_row["总派送成本"], 1349.8)
        self.assertEqual(linehaul_row["平均整车价"], 450)
        self.assertEqual(linehaul_row["每方平均价"], 10)
        self.assertEqual(linehaul_row["平均每车出库体积"], 45)
        self.assertEqual(linehaul_row["平均派送时效"], 2)
        self.assertEqual(linehaul_row["P80派送时效"], 2)

        transfer = delivery_runtime._build_transfer_report(rows).iloc[0]
        self.assertEqual(transfer["车次数"], 2)
        self.assertAlmostEqual(transfer["总出库体积"], 89.99, places=2)
        self.assertEqual(transfer["平均整车价"], 450)
        self.assertEqual(transfer["每方平均价"], 10)
        self.assertEqual(transfer["平均每车出库体积"], 45)

    def test_regular_delivery_floor_and_pallet_thresholds_filter_only_average_samples(self):
        rows = pd.DataFrame([
            {"车型装车分组": "大车地板", "出库体积": 80, "出库卡板数": 16, "派送成本": 800},
            {"车型装车分组": "大车地板", "出库体积": 79.99, "出库卡板数": 99, "派送成本": 999},
            {"车型装车分组": "大车卡板", "出库体积": 40, "出库卡板数": 12, "派送成本": 400},
            {"车型装车分组": "大车卡板", "出库体积": 39.99, "出库卡板数": 88, "派送成本": 888},
            {"车型装车分组": "小车", "出库体积": 10, "出库卡板数": 2, "派送成本": 100},
        ])

        eligible = processors.regular_delivery_average_sample_rows(rows)

        self.assertEqual(eligible.index.tolist(), [0, 2, 4])
        self.assertAlmostEqual(rows["出库体积"].sum(), 249.98, places=2)

    def test_regular_cost_report_uses_filtered_detail_for_all_average_and_p80_metrics(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-W28", "是否FTL发车": True,
                "车次号": "COST-T1",
                "车型标准值": "53尺大车", "装车类型标准值": "地板",
                "主产品类型": "FBA", "FBA仓点代码集合": "ONT8",
                "出库体积": 80, "出库卡板数": 16, "派送成本": 800,
            },
            {
                "仓库": "LA", "统计周期": "2026-W28", "是否FTL发车": True,
                "车次号": "COST-T2",
                "车型标准值": "53尺大车", "装车类型标准值": "地板",
                "主产品类型": "FBA", "FBA仓点代码集合": "ONT8",
                "出库体积": 79, "出库卡板数": 99, "派送成本": 999,
            },
        ])

        report = delivery_match_adapter.build_station_cost_report(rows)
        row = report.loc[report["指标名称"] == "FBA及FBX平台仓成本"].iloc[0]

        self.assertEqual(row["车次数"], 2)
        self.assertEqual(row["总出库体积"], 159)
        self.assertEqual(row["总出库卡板数"], 115)
        self.assertEqual(row["总派送成本"], 1799)
        self.assertEqual(row["平均整车价"], 800)
        self.assertEqual(row["P80整车价"], 800)
        self.assertEqual(row["每方平均价"], 10)
        self.assertEqual(row["平均每车出库体积"], 80)
        self.assertEqual(row["P80每车出库体积"], 80)
        self.assertEqual(row["平均每车出库卡板数"], 16)
        self.assertEqual(row["P80每车出库卡板数"], 16)

        full_load_row = report.loc[report["指标名称"] == "满载情况"].iloc[0]
        self.assertEqual(full_load_row["车次数"], 2)
        self.assertEqual(full_load_row["总出库体积"], 159)
        self.assertEqual(full_load_row["平均每车出库体积"], 80)
        self.assertEqual(full_load_row["P80每车出库体积"], 80)

        exported = tool_common.clean_for_excel_output(report, sheet_type="成本")
        exported_row = exported.loc[exported["指标名称"] == "FBA及FBX平台仓成本"].iloc[0]
        self.assertEqual(exported_row["总派送成本"], 1799)
        self.assertEqual(exported_row["平均整车价"], 800)
        self.assertEqual(exported_row["P80整车价"], 800)
        self.assertEqual(exported_row["每方平均价"], 10)
        self.assertEqual(exported_row["平均每车出库体积"], 80)
        self.assertEqual(exported_row["P80每车出库体积"], 80)
        self.assertEqual(exported_row["平均每车出库卡板数"], 16)
        self.assertEqual(exported_row["P80每车出库卡板数"], 16)

        workbook = tool_common.write_sheets_to_excel({"成本": report})
        workbook_row = pd.read_excel(workbook, sheet_name="成本").loc[
            lambda df: df["指标名称"] == "FBA及FBX平台仓成本"
        ].iloc[0]
        self.assertEqual(workbook_row["平均整车价"], 800)
        self.assertEqual(workbook_row["每方平均价"], 10)
        self.assertEqual(workbook_row["平均每车出库体积"], 80)

    def test_zy_multibatch_trip_uses_volume_share_for_vehicle_but_preserves_batch_cost(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "FTL-MULTI-1",
                "批次号": "BATCH-ONT8", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "出库体积": 40, "出库卡板数": 8, "派送成本": 350,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "车型": "53尺大车", "车型标准值": "53尺大车",
                "装车类型": "", "装车类型标准值": "未知装车类型",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "FTL-MULTI-1",
                "批次号": "BATCH-LAS1", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 220,
                "FBA/FBX": "FBA", "FBA仓点代码": "LAS1",
                "车型": "53尺大车", "车型标准值": "53尺大车",
                "装车类型": "", "装车类型标准值": "未知装车类型",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        self.assertEqual(len(cleaned), 2)
        self.assertTrue(cleaned["装车类型标准值"].eq("卡板").all())
        self.assertTrue(cleaned["派送方式"].eq("53尺大车-卡板").all())
        self.assertEqual(cleaned["整车出库体积"].unique().tolist(), [60])
        self.assertAlmostEqual(cleaned["批次车份额"].sum(), 1)
        self.assertEqual(set(cleaned["批次目的仓点"]), {"ONT8", "LAS1"})

        stage1_workbook = tool_common.write_sheets_to_excel({"派送一_清洗后批次": cleaned})
        cleaned_roundtrip = pd.read_excel(stage1_workbook, sheet_name="派送一_清洗后批次")
        delivery_runtime.bootstrap(delivery_workflow)
        metrics = delivery_workflow.process_stage2_analysis(
            cleaned_roundtrip,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            period_type="按月统计",
        )
        typed = metrics["分类型价格参考"]
        self.assertEqual(typed["仓点代码"].tolist(), ["ONT8", "LAS1"])
        self.assertTrue(typed["成本计算类型"].eq("大车卡板").all())

        ont8 = typed.loc[typed["仓点代码"] == "ONT8"].iloc[0]
        las1 = typed.loc[typed["仓点代码"] == "LAS1"].iloc[0]
        self.assertEqual(ont8["车次数"], 0.67)
        self.assertEqual(las1["车次数"], 0.33)
        self.assertEqual(ont8["细分货量方数"], 40)
        self.assertEqual(las1["细分货量方数"], 20)
        self.assertEqual(ont8["总出库卡板数"], 8)
        self.assertEqual(las1["总出库卡板数"], 4)
        self.assertEqual(ont8["总派送成本"], 350)
        self.assertEqual(las1["总派送成本"], 220)
        self.assertTrue(pd.isna(ont8["整车价格"]))
        self.assertTrue(pd.isna(las1["整车价格"]))
        self.assertEqual(ont8["每方成本"], 8.75)
        self.assertEqual(las1["每方成本"], 11)
        self.assertEqual(ont8["平均每车出库体积"], 60)
        self.assertEqual(las1["平均每车出库体积"], 60)
        self.assertEqual(ont8["平均每车出库卡板数"], 12)
        self.assertEqual(las1["平均每车出库卡板数"], 12)
        self.assertTrue(typed["仓点分摊口径"].str.contains("批次成本原值").all())

        workbook = tool_common.write_sheets_to_excel(metrics)
        exported = pd.read_excel(workbook, sheet_name="分类型价格参考")
        self.assertEqual(exported["车次数"].tolist(), [0.67, 0.33])

    def test_same_destination_multibatch_trip_is_excluded_only_from_whole_truck_cost(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "FTL-SAME-1",
                "批次号": "BATCH-A", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "出库体积": 40, "出库卡板数": 8, "派送成本": 350,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "车型": "53尺大车", "车型标准值": "53尺大车",
                "装车类型": "卡板", "装车类型标准值": "卡板",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "FTL-SAME-1",
                "批次号": "BATCH-B", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 220,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "车型": "53尺大车", "车型标准值": "53尺大车",
                "装车类型": "卡板", "装车类型标准值": "卡板",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        report = delivery_match_adapter.build_station_cost_report(cleaned.assign(
            是否FTL发车=True,
            统计周期="2026-07",
        ))
        row = report.loc[report["指标名称"] == "FBA及FBX平台仓成本"].iloc[0]

        self.assertEqual(row["车次数"], 1)
        self.assertEqual(row["总出库体积"], 60)
        self.assertEqual(row["总出库卡板数"], 12)
        self.assertEqual(row["总派送成本"], 570)
        self.assertTrue(pd.isna(row["平均整车价"]))
        self.assertTrue(pd.isna(row["P80整车价"]))
        self.assertEqual(row["每方平均价"], 9.5)
        self.assertEqual(row["平均每车出库体积"], 60)
        self.assertEqual(row["平均每车出库卡板数"], 12)

    def test_normal_cost_uses_only_single_batch_trip_for_whole_truck_price(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-07", "标准运输类型": "FTL",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-MULTI", "批次号集合": "A", "整车批次数": 2,
                "批次车份额": 2 / 3, "整车出库体积": 60, "整车出库卡板数": 12,
                "出库体积": 40, "出库卡板数": 8, "派送成本": 350,
                "车型标准值": "53尺大车", "装车类型标准值": "卡板",
                "主产品类型": "FBA", "FBA仓点代码集合": "ONT8",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "标准运输类型": "FTL",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-MULTI", "批次号集合": "B", "整车批次数": 2,
                "批次车份额": 1 / 3, "整车出库体积": 60, "整车出库卡板数": 12,
                "出库体积": 20, "出库卡板数": 4, "派送成本": 220,
                "车型标准值": "53尺大车", "装车类型标准值": "卡板",
                "主产品类型": "FBA", "FBA仓点代码集合": "ONT8",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "标准运输类型": "FTL",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-SINGLE", "批次号集合": "C", "整车批次数": 1,
                "批次车份额": 1, "整车出库体积": 60, "整车出库卡板数": 12,
                "出库体积": 60, "出库卡板数": 12, "派送成本": 600,
                "车型标准值": "53尺大车", "装车类型标准值": "卡板",
                "主产品类型": "FBA", "FBA仓点代码集合": "ONT8",
            },
        ])

        marked = processors.mark_whole_truck_cost_sample_eligibility(rows)
        self.assertTrue(marked.loc[marked["车次号"] == "T-MULTI", "是否纳入整车成本样本"].eq(False).all())
        self.assertTrue(marked.loc[marked["车次号"] == "T-SINGLE", "是否纳入整车成本样本"].eq(True).all())
        self.assertTrue(
            marked.loc[marked["车次号"] == "T-MULTI", "整车成本样本排除原因"]
            .eq("同车次含多个批次（多卸）")
            .all()
        )

        report = delivery_match_adapter.build_station_cost_report(rows)
        row = report.loc[
            (report["指标名称"] == "FBA及FBX平台仓成本")
            & (report["仓点代码"] == "ONT8")
        ].iloc[0]
        self.assertEqual(row["车次数"], 2)
        self.assertEqual(row["总出库体积"], 120)
        self.assertEqual(row["总派送成本"], 1170)
        self.assertEqual(row["平均整车价"], 600)
        self.assertEqual(row["P80整车价"], 600)
        self.assertEqual(row["每方平均价"], 9.75)
        self.assertEqual(row["平均每车出库体积"], 60)

    def test_transfer_and_linehaul_exclude_cross_group_multibatch_trip_only_from_whole_truck_price(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-NJ",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-MULTI", "批次号集合": "A", "整车批次数": 2,
                "批次车份额": 2 / 3, "整车出库体积": 90,
                "出库体积": 60, "出库卡板数": 12, "派送成本": 350, "派送时效": 2,
                "出库类型": "调拨", "调入仓库": "NJ", "业务场景": "仓间调拨",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-DAL",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-MULTI", "批次号集合": "B", "整车批次数": 2,
                "批次车份额": 1 / 3, "整车出库体积": 90,
                "出库体积": 30, "出库卡板数": 6, "派送成本": 220, "派送时效": 3,
                "出库类型": "正常", "调入仓库": "", "业务场景": "FBA派送",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-NJ",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": "T-SINGLE", "批次号集合": "C", "整车批次数": 1,
                "批次车份额": 1, "整车出库体积": 60,
                "出库体积": 60, "出库卡板数": 12, "派送成本": 600, "派送时效": 2,
                "出库类型": "调拨", "调入仓库": "NJ", "业务场景": "仓间调拨",
            },
        ])

        transfer = delivery_runtime._build_transfer_report(rows).iloc[0]
        self.assertEqual(transfer["总出库体积"], 120)
        self.assertEqual(transfer["总派送成本"], 950)
        self.assertEqual(transfer["平均整车价"], 600)
        self.assertAlmostEqual(transfer["每方平均价"], (350 / 60 + 600 / 60) / 2)

        linehaul = delivery_audit_backfill._build_linehaul_sheet(rows)
        nj = linehaul.loc[linehaul["专线线路"] == "LA-NJ"].iloc[0]
        dal = linehaul.loc[linehaul["专线线路"] == "LA-DAL"].iloc[0]
        self.assertEqual(nj["总出库体积"], 120)
        self.assertEqual(nj["总派送成本"], 950)
        self.assertEqual(nj["平均整车价"], 600)
        self.assertAlmostEqual(nj["每方平均价"], round((350 / 60 + 600 / 60) / 2, 2))
        self.assertEqual(dal["总出库体积"], 30)
        self.assertEqual(dal["总派送成本"], 220)
        self.assertTrue(pd.isna(dal["平均整车价"]))
        self.assertEqual(dal["每方平均价"], round(220 / 30, 2))

    def test_transfer_and_linehaul_append_compact_supplier_whole_truck_metrics(self):
        def row(trip, batch, supplier, cost, *, batch_count=1, share=1):
            return {
                "仓库": "LA", "统计周期": "2026-07", "专线线路": "LA-NJ",
                "是否有真实车次号": True, "是否FTL发车": True,
                "车次号": trip, "批次号集合": batch, "整车批次数": batch_count,
                "批次车份额": share, "整车出库体积": 50 if batch_count == 1 else 100,
                "出库体积": 50, "出库卡板数": 10, "派送成本": cost, "派送时效": 2,
                "派送卡车": supplier, "出库类型": "调拨", "调入仓库": "NJ",
                "业务场景": "仓间调拨",
            }

        rows = pd.DataFrame([
            row("T-A1", "A1", "Carrier A", 500),
            row("T-A2", "A2", "carrier a", 700),
            row("T-B1", "B1", "Carrier B", 900),
            row("T-BLANK", "BLANK", "", 800),
            row("T-MULTI", "M1", "Carrier B", 300, batch_count=2, share=0.5),
            row("T-MULTI", "M2", "Carrier B", 400, batch_count=2, share=0.5),
        ])

        expected_columns = [
            "发货仓", "调拨目标仓", "专线线路", "统计周期", "车次数",
            "总出库体积", "总出库卡板数", "总派送成本", "平均整车价", "每方平均价",
            "平均每车出库体积", "供应商平均整车成本", "供应商使用比例",
        ]
        transfer = delivery_runtime._build_transfer_report(rows)
        self.assertEqual(transfer.columns.tolist(), expected_columns)
        transfer_row = transfer.iloc[0]
        self.assertEqual(
            transfer_row["供应商平均整车成本"],
            "Carrier A $600.00；Carrier B $900.00",
        )
        self.assertEqual(
            transfer_row["供应商使用比例"],
            "Carrier A 50.00%；Carrier B 25.00%",
        )
        self.assertEqual(transfer_row["平均整车价"], 725)

        linehaul = delivery_audit_backfill._build_linehaul_sheet(rows)
        nj = linehaul.loc[linehaul["专线线路"] == "LA-NJ"].iloc[0]
        self.assertEqual(
            nj["供应商平均整车成本"],
            "Carrier A $600.00；Carrier B $900.00",
        )
        self.assertEqual(
            nj["供应商使用比例"],
            "Carrier A 50.00%；Carrier B 25.00%",
        )
        self.assertEqual(linehaul.columns[-2:].tolist(), ["供应商平均整车成本", "供应商使用比例"])
        self.assertNotIn("指标名称", linehaul.columns)
        self.assertNotIn("批次号集合", linehaul.columns)
        self.assertNotIn("车次号集合", linehaul.columns)

    def test_old_multidestination_rows_are_not_equal_split(self):
        rows = pd.DataFrame([{
            "仓库": "LA", "统计周期": "2026-W30", "是否FTL发车": True,
            "车型标准值": "53尺大车", "装车类型标准值": "未知装车类型",
            "主产品类型": "FBA", "FBA仓点代码集合": "ONT8,LAS1",
            "出库体积": 60, "出库卡板数": 12, "派送成本": 900,
        }])

        report = delivery_match_adapter.build_station_cost_report(rows)

        self.assertTrue(report.empty)

    def test_station_cost_source_keeps_linehaul_transfer_and_zero_cost_rows(self):
        rows = pd.DataFrame([
            {"专线线路": "LA-NJ", "批次号集合": "A", "派送成本": 500},
            {
                "专线线路": "LA-DAL", "批次号集合": "B", "派送成本": 600,
                "调入仓库": "DAL", "业务场景": "仓间调拨",
            },
            {"专线线路": "未知线路", "批次号集合": "C", "派送成本": 0},
        ])

        source = delivery_runtime._station_cost_source_rows(rows)

        self.assertEqual(source["批次号集合"].tolist(), ["A", "B", "C"])

    def test_station_cost_reference_excludes_zero_cost_volume_from_denominator(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-06", "专线线路": "LA-DAL",
                "是否FTL发车": True, "标准运输类型": "FTL", "车次号": "FTW1-T1",
                "车型标准值": "53尺大车", "装车类型标准值": "地板",
                "主产品类型": "FBA", "FBA仓点代码集合": "FTW1",
                "出库体积": 80, "出库卡板数": 40, "原始派送成本": 5000, "派送成本": 5000,
                "批次车份额": 1, "整车出库体积": 80, "整车出库卡板数": 40,
            },
            {
                "仓库": "LA", "统计周期": "2026-06", "专线线路": "LA-DAL",
                "是否FTL发车": True, "标准运输类型": "FTL", "车次号": "FTW1-T2",
                "车型标准值": "53尺大车", "装车类型标准值": "地板",
                "主产品类型": "FBA", "FBA仓点代码集合": "FTW1",
                "出库体积": 20, "出库卡板数": 10, "原始派送成本": 0, "派送成本": 200,
                "批次车份额": 1, "整车出库体积": 20, "整车出库卡板数": 10,
            },
        ])

        delivery_runtime.bootstrap(delivery_workflow)
        cost_ftl = delivery_match_adapter.build_station_cost_report(rows)
        price_reference, type_reference = delivery_match_adapter.build_cost_price_reference_reports(
            cost_ftl,
            pd.DataFrame(),
        )

        station = price_reference[price_reference["仓点代码"] == "FTW1"].iloc[0]
        self.assertEqual(station["总出库体积"], 80)
        self.assertEqual(station["总派送成本"], 5000)
        self.assertEqual(station["每方价格参考"], 62.5)

        detail = type_reference[type_reference["仓点代码"] == "FTW1"].iloc[0]
        self.assertEqual(detail["成本计算类型"], "大车地板")
        self.assertEqual(detail["细分货量方数"], 80)
        self.assertEqual(detail["每方成本"], 62.5)

    def test_trip_level_transport_rules_reclassify_mixed_and_over_60_ltl_into_ftl_cost(self):
        def detail_row(
            row_no,
            trip,
            batch,
            transport,
            volume,
            cost,
            loading="卡板",
            vehicle="53尺大车",
            delivery_truck="",
        ):
            return {
                "原始行号": row_no,
                "仓库": "LA",
                "标准运输类型": transport,
                "车次号": trip,
                "派送卡车": delivery_truck,
                "批次号": batch,
                "出库时间": "2026-07-20",
                "签收时间": "2026-07-22",
                "出库体积": volume,
                "出库卡板数": 1,
                "派送成本": cost,
                "FBA/FBX": "FBA",
                "FBA仓点代码": "ONT8",
                "标准邮编": "92551",
                "邮编前三位": "925",
                "目的州": "CA",
                "车型": vehicle,
                "装车类型": loading,
                "车型标准值": vehicle if transport == "FTL" else "不适用",
                "装车类型标准值": loading if transport == "FTL" else "散板",
            }

        detail = pd.DataFrame([
            detail_row(2, "MIX-1", "MIX-FTL", "FTL", 20, 100, loading="托盘", vehicle="53尺大车"),
            detail_row(3, "MIX-1", "MIX-LTL", "LTL", 10, 200, loading="托盘", vehicle="53尺大车"),
            detail_row(4, "OVER-1", "OVER-A", "LTL", 31, 300, loading="地板", vehicle="53尺大车"),
            detail_row(5, "OVER-1", "OVER-B", "LTL", 30, 310, loading="地板", vehicle="53尺大车"),
            detail_row(6, "EDGE-60", "EDGE-A", "LTL", 30, 150),
            detail_row(7, "EDGE-60", "EDGE-B", "LTL", 30, 150, loading="地板", vehicle="53尺大车"),
            detail_row(8, "AMZ-1", "AMZ-A", "LTL", 8, 80, delivery_truck=" Amazon   Freight "),
            detail_row(9, "AMZ-1", "AMZ-B", "LTL", 7, 70),
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)

        mixed = cleaned.loc[cleaned["车次号"] == "MIX-1"]
        self.assertEqual(len(mixed), 2)
        self.assertTrue(mixed["标准运输类型"].eq("FTL").all())
        self.assertTrue(mixed["原始运输类型集合"].eq("FTL,LTL").all())
        self.assertTrue(mixed["运输类型重判原因"].str.contains("同时含FTL和LTL").all())
        self.assertEqual(mixed["出库体积"].sum(), 30)
        self.assertEqual(set(mixed["批次号集合"]), {"MIX-FTL", "MIX-LTL"})
        self.assertAlmostEqual(mixed["批次车份额"].sum(), 1)

        over = cleaned.loc[cleaned["车次号"] == "OVER-1"]
        self.assertEqual(len(over), 2)
        self.assertTrue(over["标准运输类型"].eq("FTL").all())
        self.assertTrue(over["运输类型重判原因"].str.contains(">60CBM").all())
        self.assertEqual(over["出库体积"].sum(), 61)
        self.assertTrue(over["车型标准值"].eq("53尺大车").all())
        self.assertTrue(over["装车类型标准值"].eq("地板").all())
        self.assertTrue(over["派送方式"].eq("53尺大车-地板").all())

        edge = cleaned.loc[cleaned["车次号"] == "EDGE-60"]
        self.assertEqual(len(edge), 2)
        self.assertEqual(edge["标准运输类型"].unique().tolist(), ["LTL"])
        self.assertTrue(edge["运输类型重判原因"].fillna("").eq("").all())

        amazon_freight = cleaned.loc[cleaned["车次号"] == "AMZ-1"]
        self.assertEqual(len(amazon_freight), 2)
        self.assertTrue(amazon_freight["标准运输类型"].eq("LTL").all())
        self.assertEqual(amazon_freight["出库体积"].sum(), 15)
        self.assertEqual(set(amazon_freight["批次号集合"]), {"AMZ-A", "AMZ-B"})
        self.assertTrue(amazon_freight["运输类型重判原因"].fillna("").eq("").all())

        delivery_runtime.bootstrap(delivery_workflow)
        metrics = delivery_workflow.process_stage2_analysis(
            cleaned,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            period_type="按月统计",
        )
        price_reference = metrics["每方价格参考"]
        type_price_reference = metrics["分类型价格参考"]

        self.assertEqual(price_reference["仓点代码"].tolist(), ["ONT8"])
        self.assertEqual(price_reference.iloc[0]["总出库体积"], 166)
        self.assertEqual(price_reference.iloc[0]["总派送成本"], 1560)
        self.assertEqual(price_reference.iloc[0]["每方价格参考"], round(1560 / 166, 2))
        self.assertEqual(
            type_price_reference["成本计算类型"].tolist(),
            ["大车地板", "大车卡板", "LTL"],
        )
        ltl = type_price_reference.loc[type_price_reference["成本计算类型"] == "LTL"].iloc[0]
        self.assertEqual(ltl["总出库体积"], 75)
        self.assertEqual(ltl["总派送成本"], 450)
        self.assertEqual(ltl["目的地总出库体积"], 166)
        self.assertEqual(ltl["细分货量方数"], 75)
        self.assertEqual(ltl["每方成本"], 6)
        self.assertTrue(pd.isna(ltl["整车价格"]))
        floor = type_price_reference.loc[
            type_price_reference["成本计算类型"] == "大车地板"
        ].iloc[0]
        self.assertTrue(pd.isna(floor["平均整车价"]))
        self.assertTrue(pd.isna(floor["P80整车价"]))
        self.assertTrue(pd.isna(floor["每方平均价"]))
        self.assertTrue(pd.isna(floor["整车价格"]))
        self.assertEqual(floor["每方成本"], 13.28)

    def test_amazon_freight_carrier_does_not_reclassify_or_fabricate_trip(self):
        raw = pd.DataFrame([
            {
                "仓库": "LA", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "目的地": "Amazon-ONT8", "派送方式": "卡车派送", "运输类型": "LTL",
                "派送卡车": "AMAZON FREIGHT", "批次号": "AF-1", "出库体积": 8,
                "出库卡板数": 2, "派送成本": 80, "车型": "53尺大车", "装车类型": "卡板",
            },
            {
                "仓库": "LA", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "目的地": "Amazon-ONT8", "派送方式": "卡车派送", "运输类型": "LTL",
                "派送卡车": "amazon freight", "批次号": "AF-2", "出库体积": 7,
                "出库卡板数": 1, "派送成本": 70, "车型": "53尺大车", "装车类型": "卡板",
            },
        ])

        detail, _, _ = processors.process_delivery_stage1_from_df(raw, warehouse="LA")
        detail = delivery_stage1_adapter.repair_delivery_stage1_numeric_columns(detail)
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)

        self.assertEqual(detail["派送卡车"].str.upper().unique().tolist(), ["AMAZON FREIGHT"])
        self.assertEqual(len(cleaned), 2)
        self.assertTrue(cleaned["标准运输类型"].eq("LTL").all())
        self.assertTrue(cleaned["车次号"].fillna("").eq("").all())
        self.assertTrue(cleaned["批次车份额"].isna().all())
        self.assertEqual(set(cleaned["批次号集合"]), {"AF-1", "AF-2"})
        self.assertTrue(cleaned["运输类型重判原因"].fillna("").eq("").all())

    def test_ltl_cost_report_groups_fba_and_fbx_platform_by_station(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-W30", "标准运输类型": "LTL",
                "是否FTL发车": False, "主产品类型": "FBA",
                "车次号": "LTL-T1",
                "FBA仓点代码集合": "ONT8", "平台名称": "",
                "出库体积": 8, "出库卡板数": 2, "派送成本": 120,
            },
            {
                "仓库": "LA", "统计周期": "2026-W30", "标准运输类型": "LTL",
                "是否FTL发车": False, "主产品类型": "FBA",
                "车次号": "LTL-T2",
                "FBA仓点代码集合": "ONT8", "平台名称": "",
                "出库体积": 12, "出库卡板数": 3, "派送成本": 180,
            },
            {
                "仓库": "LA", "统计周期": "2026-W30", "标准运输类型": "LTL",
                "是否FTL发车": False, "主产品类型": "FBX",
                "车次号": "LTL-T3",
                "平台名称": "谷仓", "FBX代码集合": "16号仓",
                "平台仓配对集合": "谷仓||16号仓",
                "出库体积": 6, "出库卡板数": 1, "派送成本": 90,
            },
            {
                "仓库": "LA", "统计周期": "2026-W30", "标准运输类型": "LTL",
                "是否FTL发车": False, "主产品类型": "FBX",
                "车次号": "LTL-T4",
                "平台名称": "非平台/未知", "FBX代码集合": "",
                "出库体积": 5, "出库卡板数": 1, "派送成本": 75,
            },
            {
                "仓库": "LA", "统计周期": "2026-W30", "标准运输类型": "FTL",
                "是否FTL发车": True, "主产品类型": "FBA",
                "FBA仓点代码集合": "ONT8",
                "出库体积": 40, "出库卡板数": 8, "派送成本": 500,
            },
        ])

        report = delivery_match_adapter.build_ltl_station_cost_report(rows)

        self.assertEqual(
            report.columns.tolist(),
            [
                "指标名称", "仓库", "统计周期", "对象类型", "平台", "仓点代码", "车型装车分组",
                "总出库体积", "总出库卡板数", "总派送成本",
            ],
        )
        self.assertEqual(set(report["仓点代码"]), {"ONT8", "16号仓"})
        self.assertEqual(report["车型装车分组"].unique().tolist(), ["LTL"])

        fba = report.loc[report["仓点代码"] == "ONT8"].iloc[0]
        self.assertEqual(fba["总出库体积"], 20)
        self.assertEqual(fba["总出库卡板数"], 5)
        self.assertEqual(fba["总派送成本"], 300)

        fbx = report.loc[report["仓点代码"] == "16号仓"].iloc[0]
        self.assertEqual(fbx["对象类型"], "FBX平台仓")
        self.assertEqual(fbx["平台"], "谷仓")
        self.assertEqual(fbx["总出库体积"], 6)
        self.assertEqual(fbx["总出库卡板数"], 1)
        self.assertEqual(fbx["总派送成本"], 90)

    def test_cost_price_reference_uses_weighted_totals_and_destination_volume_order(self):
        cost_ftl = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-W30", "对象类型": "FBA", "平台": "",
                "仓点代码": "ONT8", "车型装车分组": "大车地板",
                "总出库体积": 10, "总出库卡板数": 2, "总派送成本": 100,
                "平均整车价": 88, "P80整车价": 99, "每方平均价": 12.3,
            },
            {
                "仓库": "LA", "统计周期": "2026-W31", "对象类型": "FBA", "平台": "",
                "仓点代码": "ONT8", "车型装车分组": "小车",
                "总出库体积": 30, "总出库卡板数": 6, "总派送成本": 600,
                "平均整车价": 550,
            },
            {
                "仓库": "LA", "统计周期": "2026-W31", "对象类型": "FBA", "平台": "",
                "仓点代码": "LAX9", "车型装车分组": "大车卡板",
                "总出库体积": 60, "总出库卡板数": 12, "总派送成本": 600,
                "平均整车价": 600,
            },
        ])
        cost_ltl = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-W31", "对象类型": "FBA", "平台": "",
                "仓点代码": "ONT8", "车型装车分组": "LTL",
                "总出库体积": 10, "总出库卡板数": 1, "总派送成本": 300,
            },
        ])

        price_reference, type_price_reference = (
            delivery_match_adapter.build_cost_price_reference_reports(cost_ftl, cost_ltl)
        )

        self.assertEqual(price_reference["仓点代码"].tolist(), ["LAX9", "ONT8", "ONT8"])
        self.assertEqual(price_reference["统计周期"].tolist(), ["2026-W31", "2026-W31", "2026-W30"])
        self.assertEqual(price_reference["排名"].tolist(), [1, 2, 2])
        self.assertEqual(price_reference["目的地总出库体积"].tolist(), [60, 50, 50])
        ont8_w30 = price_reference[
            (price_reference["仓点代码"] == "ONT8")
            & (price_reference["统计周期"] == "2026-W30")
        ].iloc[0]
        self.assertEqual(ont8_w30["总出库体积"], 10)
        self.assertEqual(ont8_w30["总派送成本"], 100)
        self.assertEqual(ont8_w30["每方价格参考"], 10)
        ont8_w31 = price_reference[
            (price_reference["仓点代码"] == "ONT8")
            & (price_reference["统计周期"] == "2026-W31")
        ].iloc[0]
        self.assertEqual(ont8_w31["总出库体积"], 40)
        self.assertEqual(ont8_w31["总派送成本"], 900)
        self.assertEqual(ont8_w31["每方价格参考"], 22.5)

        ont8_types = type_price_reference.loc[
            type_price_reference["仓点代码"] == "ONT8"
        ]
        self.assertEqual(
            ont8_types["成本计算类型"].tolist(),
            ["小车", "LTL", "大车地板"],
        )
        self.assertEqual(ont8_types["统计周期"].tolist(), ["2026-W31", "2026-W31", "2026-W30"])
        self.assertEqual(ont8_types["目的地总出库体积"].tolist(), [50, 50, 50])
        self.assertEqual(ont8_types["细分货量方数"].tolist(), [30, 10, 10])
        self.assertEqual(ont8_types["每方成本"].tolist(), [20, 30, 10])
        floor = ont8_types.loc[ont8_types["成本计算类型"] == "大车地板"].iloc[0]
        self.assertEqual(floor["平均整车价"], 88)
        self.assertEqual(floor["P80整车价"], 99)
        self.assertEqual(floor["每方平均价"], 12.3)
        self.assertEqual(floor["整车价格"], 88)
        self.assertEqual(floor["每方成本"], 10)
        small = ont8_types.loc[ont8_types["成本计算类型"] == "小车"].iloc[0]
        self.assertEqual(small["整车价格"], 550)
        ltl = ont8_types.loc[ont8_types["成本计算类型"] == "LTL"].iloc[0]
        self.assertTrue(pd.isna(ltl["整车价格"]))
        self.assertEqual(
            type_price_reference["仓点代码"].tolist(),
            ["LAX9", "ONT8", "ONT8", "ONT8"],
        )

    def test_golden_standard_data_uses_strict_single_batch_trip_boundaries(self):
        def golden_row(batch, trip, loading, volume, cost=500, **overrides):
            row = {
                "仓库": "LA",
                "分析批次ID": f"BATCH_{batch}",
                "批次号": batch,
                "批次号集合": batch,
                "车次号": trip,
                "是否有真实车次号": bool(trip),
                "批次数据是否有效": True,
                "标准运输类型": "FTL",
                "车型标准值": "53尺大车",
                "装车类型标准值": loading,
                "出库体积": volume,
                "出库卡板数": 10,
                "原始派送成本": cost,
                "派送成本": cost,
                "批次车份额": 1,
                "整车批次数": 1,
                "整车出库体积": volume,
                "批次目的仓点": "ONT8",
                "调拨目标仓代码": "",
                "专线线路": "LA-CHI",
                "派送区域": "CHI区域",
                "派送卡车": "Carrier A",
                "批次出库时间": pd.Timestamp("2026-07-20 10:00:00"),
                "批次创建时间": pd.Timestamp("2026-07-18 09:00:00"),
            }
            row.update(overrides)
            return row

        rows = pd.DataFrame([
            golden_row("FLOOR-80", "T-FLOOR-80", "地板", 80),
            golden_row("FLOOR-120", "T-FLOOR-120", "地板", 120),
            golden_row("PALLET-40", "T-PALLET-40", "卡板", 40),
            golden_row("PALLET-80", "T-PALLET-80", "卡板", 80),
            golden_row("FLOOR-LOW", "T-FLOOR-LOW", "地板", 79.99),
            golden_row("FLOOR-HIGH", "T-FLOOR-HIGH", "地板", 120.01),
            golden_row("PALLET-LOW", "T-PALLET-LOW", "卡板", 39.99),
            golden_row("PALLET-HIGH", "T-PALLET-HIGH", "卡板", 80.01),
            golden_row("ZERO-BASE", "T-ZERO", "地板", 90, cost=0, 派送成本=200),
            golden_row("MISSING-BASE", "T-MISSING-BASE", "地板", 90, 原始派送成本=pd.NA, 派送成本=700),
            golden_row("SMALL", "T-SMALL", "卡板", 50, 车型标准值="26尺小车"),
            golden_row("LTL", "T-LTL", "卡板", 50, 标准运输类型="LTL"),
            golden_row("NO-TRIP", "", "卡板", 50),
            golden_row("MULTI-A", "T-MULTI", "卡板", 50, 批次车份额=0.5, 整车批次数=2),
            golden_row("MULTI-B", "T-MULTI", "卡板", 50, 批次车份额=0.5, 整车批次数=2),
            golden_row("TWO-TRIPS", "T-DUP-1", "卡板", 50),
            golden_row("TWO-TRIPS", "T-DUP-2", "卡板", 50),
            golden_row("INVALID", "T-INVALID", "卡板", 50, 批次数据是否有效=False),
        ])

        result = delivery_match_adapter.build_golden_standard_batch_report(rows)

        self.assertEqual(
            result["批次号"].tolist(),
            ["FLOOR-80", "FLOOR-120", "PALLET-40", "PALLET-80"],
        )
        self.assertEqual(
            result.columns.tolist(),
            [
                "批次号", "车次号", "目的地", "调拨属于", "干线属于", "区域属于",
                "车型", "装车类型", "派送卡车", "出库体积", "出库卡板", "派送成本",
                "出库时间", "创建时间",
            ],
        )
        self.assertEqual(result["装车类型"].tolist(), ["地板", "地板", "卡板", "卡板"])
        self.assertEqual(result["干线属于"].tolist(), ["LA-CHI"] * 4)
        self.assertEqual(result["区域属于"].tolist(), ["CHI区域"] * 4)
        self.assertTrue(result["派送成本"].gt(0).all())

    def test_excel_export_preserves_leading_equals_and_colors_suppliers(self):
        golden = pd.DataFrame([
            {"批次号": "A", "派送卡车": "Carrier A", "派送成本": 500},
            {"批次号": "B", "派送卡车": "Carrier B", "派送成本": 600},
            {"批次号": "C", "派送卡车": "carrier a", "派送成本": 700},
        ])
        detail = pd.DataFrame([
            {"批次号": "A", "同车次备注集合": "=10758801=349件；按PO打板"},
        ])

        output = tool_common.write_sheets_to_excel({
            "黄金标准数据": golden,
            "派送二_匹配后批次数据": detail,
        })
        workbook = load_workbook(output, data_only=False)
        detail_ws = workbook["派送二_匹配后批次数据"]
        self.assertEqual(detail_ws["B2"].data_type, "s")
        self.assertEqual(detail_ws["B2"].value, "=10758801=349件；按PO打板")

        golden_ws = workbook["黄金标准数据"]
        supplier_col = next(cell.column for cell in golden_ws[1] if cell.value == "派送卡车")
        fills = [golden_ws.cell(row=row, column=supplier_col).fill.fgColor.rgb for row in [2, 3, 4]]
        self.assertEqual(fills[0], fills[2])
        self.assertNotEqual(fills[0], fills[1])
        self.assertEqual(golden_ws.cell(row=2, column=supplier_col).fill.fill_type, "solid")
        workbook.close()

    def test_monthly_price_reference_does_not_merge_june_and_july(self):
        cost_ftl = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-06", "对象类型": "FBA", "平台": "FBA",
                "仓点代码": "ONT8", "车型装车分组": "大车卡板",
                "车次数": 1, "总出库体积": 40, "总出库卡板数": 8,
                "总派送成本": 400, "平均整车价": 400,
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "对象类型": "FBA", "平台": "FBA",
                "仓点代码": "ONT8", "车型装车分组": "大车卡板",
                "车次数": 1, "总出库体积": 20, "总出库卡板数": 4,
                "总派送成本": 300, "平均整车价": 300,
            },
        ])

        price_reference, type_reference = delivery_match_adapter.build_cost_price_reference_reports(
            cost_ftl,
            pd.DataFrame(),
        )

        self.assertEqual(price_reference["统计周期"].tolist(), ["2026-07", "2026-06"])
        self.assertEqual(price_reference["排名"].tolist(), [1, 1])
        self.assertEqual(price_reference["目的地总出库体积"].tolist(), [60, 60])
        self.assertEqual(price_reference["总出库体积"].tolist(), [20, 40])
        self.assertEqual(price_reference["每方价格参考"].tolist(), [15, 10])
        self.assertEqual(type_reference["统计周期"].tolist(), ["2026-07", "2026-06"])
        self.assertEqual(type_reference["目的地总出库体积"].tolist(), [60, 60])

    def test_stage2_workbook_combines_cost_price_reference_sheets_end_to_end(self):
        source = pd.DataFrame([
            {
                "仓库": "LA", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "目的地": "Amazon-ONT8", "派送方式": "卡车派送", "运输类型": "LTL",
                "车次号": "LTL-FBA-T1",
                "创建时间": "2026-07-18",
                "批次号": "LTL-FBA-1", "出库体积": 8, "出库卡板数": 2, "派送成本": 120,
            },
            {
                "仓库": "LA", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "目的地": "谷仓16号仓", "派送方式": "卡车派送", "运输类型": "LTL",
                "车次号": "LTL-FBX-T1",
                "创建时间": "2026-07-18",
                "批次号": "LTL-FBX-1", "出库体积": 6, "出库卡板数": 1, "派送成本": 90,
            },
            {
                "仓库": "LA", "出库时间": "2026-07-20", "签收时间": "2026-07-22",
                "目的地": "Amazon-ONT8", "派送方式": "卡车派送", "运输类型": "FTL",
                "车型": "53尺大车", "装车类型": "卡板", "车次号": "FTL-1",
                "创建时间": "2026-07-18", "派送卡车": "Carrier A",
                "批次号": "FTL-FBA-1", "出库体积": 40, "出库卡板数": 10, "派送成本": 300,
            },
        ])

        delivery_runtime.bootstrap(delivery_workflow)
        detail, _, _ = processors.process_delivery_stage1_from_df(source, "LA")
        detail = delivery_stage1_adapter.repair_delivery_stage1_numeric_columns(detail)
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        metrics = delivery_workflow.process_stage2_analysis(
            cleaned,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            period_type="按月统计",
        )

        self.assertIn("每方价格参考", metrics)
        self.assertIn("分类型价格参考", metrics)
        self.assertIn("黄金标准数据", metrics)
        self.assertNotIn("成本FTL", metrics)
        self.assertNotIn("成本LTL", metrics)

        golden = metrics["黄金标准数据"]
        self.assertEqual(golden["批次号"].tolist(), ["FTL-FBA-1"])
        self.assertEqual(golden.iloc[0]["装车类型"], "卡板")
        self.assertEqual(golden.iloc[0]["创建时间"], pd.Timestamp("2026-07-18"))

        price_reference = metrics["每方价格参考"]
        self.assertEqual(price_reference["仓点代码"].tolist(), ["ONT8", "16号仓"])
        self.assertEqual(price_reference["排名"].tolist(), [1, 2])
        ont8 = price_reference.loc[price_reference["仓点代码"] == "ONT8"].iloc[0]
        self.assertEqual(ont8["总出库体积"], 48)
        self.assertEqual(ont8["总出库卡板数"], 12)
        self.assertEqual(ont8["总派送成本"], 420)
        self.assertEqual(ont8["每方价格参考"], round(420 / 48, 2))

        type_price_reference = metrics["分类型价格参考"]
        self.assertEqual(
            type_price_reference[["仓点代码", "成本计算类型"]].values.tolist(),
            [["ONT8", "大车卡板"], ["ONT8", "LTL"], ["16号仓", "LTL"]],
        )
        car_count_index = type_price_reference.columns.tolist().index("车次数")
        self.assertEqual(
            type_price_reference.columns[car_count_index:car_count_index + 4].tolist(),
            ["车次数", "细分货量方数", "整车价格", "每方成本"],
        )
        self.assertEqual(type_price_reference["总出库体积"].sum(), 54)
        self.assertEqual(type_price_reference["总出库卡板数"].sum(), 13)
        self.assertEqual(type_price_reference["总派送成本"].sum(), 510)
        ftl_row = type_price_reference.loc[
            type_price_reference["成本计算类型"] == "大车卡板"
        ].iloc[0]
        self.assertIn("平均整车价", type_price_reference.columns)
        self.assertEqual(ftl_row["平均整车价"], 300)
        self.assertEqual(ftl_row["整车价格"], 300)
        self.assertEqual(ftl_row["细分货量方数"], 40)
        self.assertEqual(ftl_row["每方成本"], 7.5)
        self.assertTrue(
            type_price_reference.loc[
                type_price_reference["成本计算类型"] == "LTL",
                "整车价格",
            ].isna().all()
        )

        workbook = tool_common.write_sheets_to_excel(metrics)
        xls = pd.ExcelFile(workbook)
        self.assertIn("每方价格参考", xls.sheet_names)
        self.assertIn("分类型价格参考", xls.sheet_names)
        self.assertIn("黄金标准数据", xls.sheet_names)
        workbook_golden = pd.read_excel(workbook, sheet_name="黄金标准数据")
        self.assertEqual(workbook_golden["批次号"].tolist(), ["FTL-FBA-1"])
        self.assertIn("派送二_匹配后批次数据", xls.sheet_names)
        self.assertIn("派送二_车次汇总核对", xls.sheet_names)
        self.assertNotIn("成本FTL", xls.sheet_names)
        self.assertNotIn("成本LTL", xls.sheet_names)

    def test_added_fba_references_fill_zip_state_and_station_code(self):
        cases = [
            ("MCI4", "10501 NW 136th St, Kansas City, MO 64153", "64153", "MO"),
            ("CMH7", "1245 Beech Rd SW, New Albany, OH 43054", "43054", "OH"),
        ]
        for code, address, zip_code, state in cases:
            with self.subTest(code=code):
                reference = delivery_reference.match_fba_reference(f"Amazon-{code}")

                self.assertEqual(reference["代码"], code)
                self.assertEqual(reference["邮编"], zip_code)
                self.assertEqual(reference["州"], state)
                self.assertEqual(delivery_reference.FBA_REFERENCE_MAP[code]["地址"], address)

                rows = pd.DataFrame([{
                    "仓库": "LA",
                    "系统产品类型": "FBA",
                    "目的地": f"Amazon-{code}",
                    "修正后目的地": f"Amazon-{code}",
                    "FBA仓点代码": code,
                    "邮编是否有效": False,
                }])
                matched = delivery_reference.apply_delivery_reference_memory(rows).iloc[0]

                self.assertEqual(matched["标准邮编"], zip_code)
                self.assertEqual(matched["邮编前三位"], zip_code[:3])
                self.assertEqual(matched["目的州"], state)
                self.assertEqual(matched["规则匹配代码"], code)
                self.assertFalse(bool(matched["目的地邮编待补充"]))

    def test_zz_fbx_platform_and_text_code_flow_through_stage1_and_stage2_outputs(self):
        source = pd.DataFrame([{
            "仓库": "LA",
            "出库时间": "2026-07-20",
            "签收时间": "2026-07-22",
            "目的地": "谷仓16号仓",
            "派送方式": "卡车派送",
            "运输类型": "FTL",
            "车型": "53尺大车",
            "装车类型": "卡板",
            "车次号": "T-FBX-1",
            "批次号": "B-FBX-1",
            "出库体积": 30,
            "出库卡板数": 10,
            "派送成本": 300,
        }])

        detail, _, _ = processors.process_delivery_stage1_from_df(source, "LA")
        detail_row = detail.iloc[0]
        self.assertEqual(detail_row["系统产品类型"], "FBX平台仓")
        self.assertEqual(detail_row["平台名称"], "谷仓")
        self.assertEqual(detail_row["FBX代码"], "16号仓")

        detail = delivery_stage1_adapter.repair_delivery_stage1_numeric_columns(detail)
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        cleaned_row = cleaned.iloc[0]
        self.assertEqual(cleaned_row["平台名称"], "谷仓")
        self.assertEqual(cleaned_row["FBX代码集合"], "16号仓")
        self.assertEqual(cleaned_row["平台仓代码集合"], "16号仓")
        self.assertEqual(cleaned_row["平台仓配对集合"], "谷仓||16号仓")

        delivery_runtime.bootstrap(delivery_workflow)
        matched = delivery_workflow.prepare_stage2_for_report(
            cleaned,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            "按周统计",
        )
        self.assertEqual(matched.iloc[0]["FBX代码集合"], "16号仓")

        report = delivery_match_adapter.build_fbx_platform_warehouse_sheet(matched)
        report_row = report.iloc[0]
        self.assertEqual(report_row["平台仓"], "谷仓")
        self.assertEqual(report_row["FBX代码"], "16号仓")

        workbook = tool_common.write_sheets_to_excel({"FBX平台仓货量": report})
        exported_row = pd.read_excel(workbook, sheet_name="FBX平台仓货量").iloc[0]
        self.assertEqual(exported_row["平台仓"], "谷仓")
        self.assertEqual(exported_row["FBX代码"], "16号仓")

    def test_batch_volume_weighted_time_and_p80_exclude_ltl_missing_trip_and_marked_rows(self):
        rows = pd.DataFrame([
            {
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "FTL", "是否有真实车次号": True,
                "是否FTL发车": True, "车次号": "T1", "批次车份额": 1,
                "出库体积": 10, "派送时效": 1, "主产品类型": "FBA",
                "FBA出库体积": 10, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "卡板",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "FTL", "是否有真实车次号": True,
                "是否FTL发车": True, "车次号": "T2", "批次车份额": 1,
                "出库体积": 30, "派送时效": 3, "主产品类型": "FBA",
                "FBA出库体积": 30, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "卡板",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "LTL", "是否有真实车次号": True,
                "是否FTL发车": False, "车次号": "LTL1", "批次车份额": pd.NA,
                "出库体积": 100, "派送时效": 2, "主产品类型": "FBA",
                "FBA出库体积": 100, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "散板",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "FTL", "是否有真实车次号": False,
                "是否FTL发车": False, "车次号": "", "批次车份额": pd.NA,
                "出库体积": 100, "派送时效": 1, "主产品类型": "FBA",
                "FBA出库体积": 100, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "卡板",
            },
            {
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "FTL", "是否有真实车次号": True,
                "是否FTL发车": True, "车次号": "T3", "批次车份额": 1,
                "出库体积": 50, "派送时效": 2, "备注": "外仓两卸", "主产品类型": "FBA",
                "FBA出库体积": 50, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "卡板",
            },
        ])

        rows["平台名称"] = ""
        report = delivery_workflow.build_sheet1_volume_dispatch_time_report(rows)
        timing = report[
            (report["指标名称"] == "目的仓点派送时效")
            & (report["维度值"] == "ONT8")
        ].iloc[0]
        self.assertEqual(timing["平均派送时效"], 2.5)
        self.assertEqual(timing["P80派送时效"], 3)
        self.assertEqual(timing["有效时效批次数"], 2)
        self.assertEqual(timing["有效时效方数"], 40)

    def test_destination_dispatch_rounds_only_after_exact_share_sum(self):
        rows = []
        for index, share in enumerate([1, 1, 1, 1, 1, 1, 0.5], start=1):
            rows.append({
                "仓库": "LA", "统计周期": "2026-07", "派送区域": "Local",
                "批次目的仓点": "ONT8", "标准运输类型": "FTL", "是否有真实车次号": True,
                "是否FTL发车": True, "车次号": f"T{index}", "批次车份额": share,
                "出库体积": 30, "派送时效": 2, "主产品类型": "FBA",
                "FBA出库体积": 30, "FBX出库体积": 0, "FBA仓点代码集合": "ONT8", "装车类型标准值": "卡板",
            })
        fixture = pd.DataFrame(rows)
        fixture["平台名称"] = ""
        report = delivery_workflow.build_sheet1_volume_dispatch_time_report(fixture)
        station = report[
            (report["指标名称"] == "目的仓点发车数")
            & (report["维度值"] == "ONT8")
        ].iloc[0]
        self.assertEqual(station["精确车份额"], 6.5)
        self.assertEqual(station["数值"], 7)

    def test_invalid_destination_or_volume_is_rejected_without_equal_split(self):
        multiple_destinations = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1",
                "批次号": "B1", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 10, "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1",
                "批次号": "B1", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 10, "FBA/FBX": "FBA", "FBA仓点代码": "LAS1",
            },
        ])
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(multiple_destinations)
        self.assertTrue(cleaned.empty)
        invalid = pd.DataFrame(cleaned.attrs["batch_invalid_records"])
        self.assertIn("多个FBA目的仓点", invalid.iloc[0]["批次无效原因"])

        missing_volume = multiple_destinations.iloc[[0]].copy()
        missing_volume["出库体积"] = 0
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(missing_volume)
        self.assertTrue(cleaned.empty)
        invalid = pd.DataFrame(cleaned.attrs["batch_invalid_records"])
        self.assertIn("无效出库体积", invalid.iloc[0]["批次无效原因"])

    def test_stage1_keeps_canonical_volume_and_pallet_headers_before_batch_validation(self):
        """真实导出表头进入批次有效性校验前，不得被通用别名二次改名成0。"""
        source = pd.DataFrame([{
            "装车类型": "卡板",
            "批次号": "PC2606010066",
            "出库体积": 73.2577,
            "目的地": "Amazon-TOL3",
            "派送卡车": "UPS",
            "车次号": "CC260601000053",
            "出库时间": "2026-06-01 19:16:47",
            "签收时间": "",
            "出库卡板数": 21,
            "派送方式": "卡车派送",
            "出库类型": "派送",
            "批次状态": "",
            "派送成本": 350,
            "运输类型": "FTL",
            "车型": "53尺大车",
        }])

        delivery_runtime.bootstrap(delivery_workflow)
        cleaned, invalid, zip_audit, detail = (
            delivery_workflow.process_stage1_raw_files_to_cleaned_batches(
                [("列表数据.xlsx", source)],
                "LA",
            )
        )

        self.assertEqual(len(cleaned), 1)
        self.assertTrue(invalid.empty)
        self.assertEqual(detail.iloc[0]["出库体积"], 73.2577)
        self.assertEqual(detail.iloc[0]["出库卡板数"], 21)
        self.assertEqual(cleaned.iloc[0]["出库体积"], 73.2577)
        self.assertEqual(cleaned.iloc[0]["出库卡板数"], 21)
        self.assertEqual(cleaned.iloc[0]["派送卡车"], "UPS")
        self.assertNotIn("缺少或无效出库体积", str(invalid.get("无效批次剔除原因", "")))
        self.assertIsInstance(zip_audit, pd.DataFrame)

    def test_batch_delivery_truck_conflict_keeps_volume_without_supplier_assignment(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1",
                "批次号": "B1", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 20, "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "派送卡车": "JTeam INC",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T1",
                "批次号": "B1", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 30, "FBA/FBX": "FBA", "FBA仓点代码": "ONT8",
                "派送卡车": "GN Trucking",
            },
        ])

        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)

        self.assertEqual(cleaned.iloc[0]["出库体积"], 50)
        self.assertEqual(cleaned.iloc[0]["派送卡车"], "")
        self.assertIn("多个派送卡车", cleaned.iloc[0]["供应商审核"])
        self.assertIn("计入占比分母", cleaned.iloc[0]["供应商审核"])

    def test_stage2_destination_volume_tables_append_delivery_truck_volume_share(self):
        rows = pd.DataFrame([
            {"仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 20, "FBA仓点代码集合": "ONT8", "派送卡车": "JTeam INC"},
            {"仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 35, "FBA仓点代码集合": "ONT8", "派送卡车": "jteam inc"},
            {"仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 25, "FBA仓点代码集合": "ONT8", "派送卡车": "GN Trucking"},
            {"仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 10, "FBA仓点代码集合": "ONT8", "派送卡车": "AMAZON FREIGHT"},
            {"仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 10, "FBA仓点代码集合": "ONT8", "派送卡车": ""},
            {"仓库": "LA", "统计周期": "2026-07", "FBA出库体积": 40, "FBA仓点代码集合": "ONT8", "派送卡车": "July Carrier"},
            {"仓库": "NJ", "统计周期": "2026-06", "FBA出库体积": 30, "FBA仓点代码集合": "ONT8", "派送卡车": "NJ Carrier"},
            {
                "仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 0,
                "FBX出库体积": 30, "平台名称": "谷仓", "FBX代码集合": "16号仓",
                "平台仓配对集合": "谷仓||16号仓", "派送卡车": "JTeam INC",
            },
            {
                "仓库": "LA", "统计周期": "2026-06", "FBA出库体积": 0,
                "FBX出库体积": 20, "平台名称": "谷仓", "FBX代码集合": "16号仓",
                "平台仓配对集合": "谷仓||16号仓", "派送卡车": "",
            },
        ]).fillna({
            "FBA出库体积": 0, "FBX出库体积": 0, "平台名称": "",
            "FBX代码集合": "", "平台仓配对集合": "",
        })

        fba = delivery_match_adapter.build_fba_rank_sheet(rows)
        fba_june_la = fba[
            (fba["仓库"] == "LA")
            & (fba["统计周期"] == "2026-06")
            & (fba["FBA仓点"] == "ONT8")
        ].iloc[0]
        self.assertEqual(
            fba_june_la["派送卡车使用比例（>10%）"],
            "JTeam INC 55.00%；GN Trucking 25.00%",
        )
        self.assertNotIn("AMAZON FREIGHT", fba_june_la["派送卡车使用比例（>10%）"])
        self.assertEqual(
            fba[(fba["仓库"] == "LA") & (fba["统计周期"] == "2026-07")].iloc[0]["派送卡车使用比例（>10%）"],
            "July Carrier 100.00%",
        )
        self.assertEqual(
            fba[(fba["仓库"] == "NJ") & (fba["统计周期"] == "2026-06")].iloc[0]["派送卡车使用比例（>10%）"],
            "NJ Carrier 100.00%",
        )
        self.assertEqual(fba.columns[-1], "派送卡车使用比例（>10%）")

        fbx = delivery_match_adapter.build_fbx_platform_warehouse_sheet(rows)
        self.assertEqual(fbx.iloc[0]["派送卡车使用比例（>10%）"], "JTeam INC 60.00%")
        self.assertEqual(fbx.columns[-1], "派送卡车使用比例（>10%）")

    def test_missing_trip_keeps_batch_cost_but_excludes_dispatch_time_and_truck_price(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "",
                "批次号": "NO-TRIP-FTL", "创建时间": "2026-06-30", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 20, "出库卡板数": 4, "派送成本": 200,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8", "标准邮编": "92551",
                "车型": "53尺大车", "装车类型": "卡板",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "LTL", "车次号": "",
                "批次号": "NO-TRIP-LTL", "创建时间": "2026-06-29", "出库时间": "2026-07-01", "签收时间": "2026-07-02",
                "出库体积": 10, "出库卡板数": 2, "派送成本": 120,
                "FBA/FBX": "FBA", "FBA仓点代码": "ONT8", "标准邮编": "92551",
                "车型": "", "装车类型": "散板",
            },
        ])
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        self.assertEqual(cleaned.iloc[0]["出库体积"], 20)
        self.assertTrue(cleaned["批次车份额"].isna().all())

        delivery_runtime.bootstrap(delivery_workflow)
        metrics = delivery_workflow.process_stage2_analysis(
            cleaned,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            "按月统计",
        )
        volume = metrics["FBA货量排行"].iloc[0]
        self.assertEqual(volume["统计周期"], "2026-07")
        self.assertEqual(volume["出库体积"], 30)
        cost_rows = metrics["分类型价格参考"].set_index("成本计算类型")
        self.assertEqual(set(cost_rows.index), {"大车卡板", "LTL"})
        self.assertTrue(cost_rows["统计周期"].eq("2026-07").all())
        self.assertEqual(cost_rows.loc["大车卡板", "细分货量方数"], 20)
        self.assertEqual(cost_rows.loc["大车卡板", "总派送成本"], 200)
        self.assertTrue(pd.isna(cost_rows.loc["大车卡板", "整车价格"]))
        self.assertEqual(cost_rows.loc["大车卡板", "车次数"], 0)
        self.assertEqual(cost_rows.loc["LTL", "细分货量方数"], 10)
        self.assertEqual(cost_rows.loc["LTL", "总派送成本"], 120)
        self.assertTrue(pd.isna(cost_rows.loc["LTL", "整车价格"]))
        self.assertFalse((metrics["发车量"]["指标名称"] == "目的仓点发车数").any())
        timing = metrics["派送时效"]
        self.assertTrue(timing.empty or "平均派送时效" not in timing.columns or timing["平均派送时效"].isna().all())
        self.assertTrue(timing.empty or "P80派送时效" not in timing.columns or timing["P80派送时效"].isna().all())

    def test_original_file_period_uses_outbound_range_for_operations_and_cost(self):
        delivery_runtime.bootstrap(delivery_workflow)
        rows = pd.DataFrame([
            {"批次出库时间": "2026-07-01", "批次创建时间": "2026-06-05"},
            {"批次出库时间": "2026-07-20", "批次创建时间": "2026-06-18"},
        ])

        result = delivery_workflow.add_analysis_period(rows, "按原文件时间范围")

        self.assertTrue(result["统计周期"].eq("2026-07-01 ~ 2026-07-20").all())
        self.assertTrue(result["成本统计周期"].eq("2026-07-01 ~ 2026-07-20").all())

    def test_weekly_cost_period_uses_outbound_time_not_creation_time(self):
        rows = pd.DataFrame([{
            "批次出库时间": "2026-08-05",
            "批次创建时间": "2026-07-01",
        }])

        result = delivery_workflow.add_analysis_period(rows, "按周统计")

        expected_period = "2026-08-03 ~ 2026-08-09"
        self.assertEqual(result.iloc[0]["统计周期"], expected_period)
        self.assertEqual(result.iloc[0]["成本统计周期"], expected_period)

    def test_same_trip_batches_keep_their_own_month_period(self):
        detail = pd.DataFrame([
            {
                "原始行号": 2, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T-PERIOD",
                "批次号": "P1", "出库时间": "2026-06-30", "签收时间": "2026-07-02",
                "出库体积": 40, "FBA/FBX": "FBA", "FBA仓点代码": "ONT8", "标准邮编": "92551",
            },
            {
                "原始行号": 3, "仓库": "LA", "标准运输类型": "FTL", "车次号": "T-PERIOD",
                "批次号": "P2", "出库时间": "2026-07-01", "签收时间": "2026-07-03",
                "出库体积": 20, "FBA/FBX": "FBA", "FBA仓点代码": "LAS1", "标准邮编": "89115",
            },
        ])
        cleaned = delivery_workflow.build_cleaned_batches_from_detail(detail)
        delivery_runtime.bootstrap(delivery_workflow)
        matched = delivery_workflow.prepare_stage2_for_report(
            cleaned,
            pd.DataFrame(columns=["批次号", "标准邮编"]),
            "按月统计",
        )
        self.assertEqual(set(matched["统计周期"]), {"2026-06", "2026-07"})
        self.assertAlmostEqual(matched["批次车份额"].sum(), 1)


if __name__ == "__main__":
    unittest.main()
