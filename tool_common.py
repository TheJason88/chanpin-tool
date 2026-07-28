import json
import re
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.utils import get_column_letter

import processors


FIELD_ALIASES = {
    "仓库": ["仓库", "入库仓库", "发货仓", "出库仓", "仓点", "warehouse", "Warehouse"],
    "批次号": ["批次号", "批次", "批次编号", "Batch", "Batch No", "batch_no"],
    "车次号": ["车次号", "派送车次号", "车次", "Trip", "Trip No"],
    "出库体积": ["出库体积", "出库方数", "方数", "体积", "CBM", "cbm", "Volume", "volume", "立方数", "出库CBM"],
    "出库卡板数": ["出库卡板数", "出库板数", "卡板数", "板数", "托盘数", "Pallets", "pallets", "Pallet", "pallet"],
    "派送成本": ["派送成本", "成本", "派送费用", "DeliveryCost", "Delivery Cost", "Cost", "cost"],
    "标准邮编": ["标准邮编", "目的地邮编", "邮编", "ZIP", "Zip", "zipcode", "ZipCode", "PostalCode", "Postal Code"],
    "目的州": ["目的州", "省/州", "州", "到达州", "目的地州", "State", "Destination State"],
    "平台名称": ["平台名称", "平台仓", "平台", "渠道", "客户平台"],
    "平台仓代码": ["平台仓代码", "FBX代码", "仓库代码", "仓库Code", "仓点代码", "目的仓代码", "Warehouse Code"],
}

TRANSFER_WAREHOUSE_INFO = {
    "LA": {
        "display": "LA盈仓",
        "zip": "91708",
        "zip3": "917",
        "state": "CA",
        "line": "LA",
        "keywords": ["LA", "美西", "洛杉矶", "CHINO", "SAN ANTONIO"],
    },
    "NJ": {
        "display": "新泽西盈仓",
        "zip": "08857",
        "zip3": "088",
        "state": "NJ",
        "line": "LA-NJ",
        "keywords": ["NJ", "新泽西", "NEW JERSEY", "OLD BRIDGE", "JAKE BROWN"],
    },
    "SAV": {
        "display": "萨凡纳盈仓",
        "zip": "31408",
        "zip3": "314",
        "state": "GA",
        "line": "LA-SAV",
        "keywords": ["SAV", "萨凡纳", "SAVANNAH", "GARDEN CITY", "PROSPERITY"],
    },
    "DAL": {
        "display": "达拉斯盈仓",
        "zip": "75180",
        "zip3": "751",
        "state": "TX",
        "line": "LA-DAL",
        "keywords": ["DAL", "达拉斯", "DALLAS", "BALCH SPRINGS", "PEACHTREE"],
    },
}

INTEGER_OUTPUT_COLUMNS = ["排名", "发车数", "派送数", "出库卡板数"]
DECIMAL_OUTPUT_COLUMNS = [
    "数值", "占比", "出库体积", "FBA出库体积", "FBX出库体积", "派送成本", "派送时效",
    "总出库体积", "总出库卡板数", "总派送成本", "平均整车价", "P80整车价", "每方平均价",
    "平均每车出库体积", "P80每车出库体积", "平均每车出库卡板数", "P80每车出库卡板数",
    "平均派送时效", "P80派送时效", "每方价格参考", "目的地总出库体积",
    "细分货量方数", "整车价格", "每方成本",
]

TEXT_COLUMN_KEYWORDS = ["邮编", "ZIP", "zip", "批次号", "车次号"]

PLATFORM_DISPLAY_MAP = {
    "walmart": "Walmart",
    "wal-mart": "Walmart",
    "tiktok": "TikTok",
    "tik-tok": "TikTok",
    "temu": "TEMU",
    "shein": "SHEIN",
    "newegg": "Newegg",
    "wayfair": "Wayfair",
    "amazon": "Amazon",
    "4px": "4PX",
}

PLATFORM_COLUMNS = ["平台仓", "平台", "平台名称", "补充平台名称"]
CODE_COLUMNS = [
    "FBX代码", "FBX代码集合", "平台仓代码", "仓点代码", "仓库代码", "平台仓代码集合", "补充平台仓代码",
    "FBA仓点", "FBA仓点代码", "FBA仓点代码集合", "仓点",
]
PAIR_COLUMNS = ["平台仓配对集合", "补充平台仓配对"]


def clean_header(value):
    return processors.clean_col_name(value)


def normalize_columns(df):
    out = df.copy()
    out.columns = [clean_header(c) for c in out.columns]
    return out


def alias_candidates(name):
    return FIELD_ALIASES.get(name, [name])


def find_first_column(df, canonical_name):
    clean_to_originals = {}
    for col in df.columns:
        clean_to_originals.setdefault(clean_header(col), []).append(col)
    for candidate in alias_candidates(canonical_name):
        clean = clean_header(candidate)
        originals = clean_to_originals.get(clean, [])
        if originals:
            return originals[0]
    return None


def normalize_numeric_series(series):
    text = series.astype(str).str.replace(",", "", regex=False).str.replace("$", "", regex=False)
    text = text.str.extract(r"(-?\d+(?:\.\d+)?)", expand=False)
    return pd.to_numeric(text, errors="coerce")


def best_numeric_alias(df, canonical_name):
    candidates = []
    clean_to_originals = {}
    for col in df.columns:
        clean_to_originals.setdefault(clean_header(col), []).append(col)
    for alias in alias_candidates(canonical_name):
        for col in clean_to_originals.get(clean_header(alias), []):
            if col not in candidates:
                candidates.append(col)
    if not candidates:
        return None
    scored = []
    for col in candidates:
        value_sum = float(normalize_numeric_series(df[col]).fillna(0).sum())
        scored.append((col, value_sum))
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored[0][0]


def repair_numeric_columns(df, columns=("出库体积", "出库卡板数", "派送成本")):
    out = normalize_columns(df)
    for target in columns:
        best_col = best_numeric_alias(out, target)
        if best_col is None:
            if target not in out.columns:
                out[target] = 0.0
            else:
                out[target] = normalize_numeric_series(out[target]).fillna(0).astype(float)
            continue
        target_sum = float(normalize_numeric_series(out[target]).fillna(0).sum()) if target in out.columns else 0.0
        best_sum = float(normalize_numeric_series(out[best_col]).fillna(0).sum())
        if target not in out.columns or target_sum <= 0 < best_sum:
            out[target] = normalize_numeric_series(out[best_col]).fillna(0).astype(float)
        else:
            out[target] = normalize_numeric_series(out[target]).fillna(0).astype(float)
    return out


def normalize_zip_value(value):
    zip_code, fix, valid, reason = processors.normalize_zip_value(value)
    return zip_code, fix, bool(valid), reason


def normalize_boolean_series(series):
    true_values = {"true", "1", "是", "yes", "y", "t"}
    return series.astype(str).str.strip().str.lower().isin(true_values)


def ensure_object_df(df):
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].astype(object)
    return out


def split_values(value):
    if processors.is_blank(value):
        return []
    parts = re.split(r"[,，;；/\s]+", str(value))
    return [p.strip() for p in parts if p.strip() and p.strip().lower() not in ["nan", "none", "null", "false", "0"]]


def _transfer_keyword_matches(upper_text, keyword):
    key = str(keyword).upper().strip()
    if not key:
        return False
    if key.isalpha() and len(key) <= 3:
        return bool(
            re.search(
                rf"(?<![A-Z0-9]){re.escape(key)}(?![A-Z0-9])",
                upper_text,
            )
        )
    return key in upper_text


def infer_transfer_target_from_text(text):
    upper = str(text).upper()
    for target, info in TRANSFER_WAREHOUSE_INFO.items():
        for keyword in info["keywords"]:
            if _transfer_keyword_matches(upper, keyword):
                return target, info
    return "", None


def infer_transfer_targets_from_text(text):
    """返回文本中命中的全部调拨目标仓，供同车次目标冲突审核使用。"""
    upper = str(text).upper()
    targets = []
    for target, info in TRANSFER_WAREHOUSE_INFO.items():
        if any(
            _transfer_keyword_matches(upper, keyword)
            for keyword in info["keywords"]
        ):
            targets.append(target)
    return targets


def infer_transfer_targets_from_row(row):
    """调入仓库优先；当前字段命中后不再让低优先级旧目的地制造冲突。"""
    preferred_cols = [
        "调入仓库", "出库类型", "业务场景", "实际目的地", "修正后目的地",
        "目的地", "备注", "车次号", "批次号集合",
    ]
    for col in preferred_cols:
        if col not in row.index:
            continue
        targets = infer_transfer_targets_from_text(row.get(col, ""))
        if targets:
            return targets
    return []


def infer_transfer_target_from_row(row):
    targets = infer_transfer_targets_from_row(row)
    if not targets:
        return "", None
    target = targets[0]
    return target, TRANSFER_WAREHOUSE_INFO[target]


TRANSFER_AUDIT_COLUMN = "调拨覆盖审核"
TRANSFER_ERROR_PREFIX = "异常："
TRANSFER_SEMANTIC_COLUMNS = ["出库类型", "业务场景", "调入仓库", "备注"]
TRANSFER_DESTINATION_TEXT_COLUMNS = [
    "实际目的地", "修正后目的地", "目的地", "标准地址", "调入仓库",
]
TRANSFER_RAW_CODE_COLUMNS = ["FBA仓点代码", "FBX代码", "平台仓代码"]
TRANSFER_BATCH_CODE_COLUMNS = [
    "FBA仓点代码集合", "FBX代码集合", "平台仓代码集合", "平台仓配对集合",
]


def _clean_transfer_text(value):
    if processors.is_blank(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "null", "<na>"} else text


def transfer_row_has_semantics(row):
    """调入仓库非空，或业务字段明确含调拨/仓间/调入，即视为调拨。"""
    if _clean_transfer_text(row.get("调入仓库", "")):
        return True
    text = " ".join(
        _clean_transfer_text(row.get(col, ""))
        for col in TRANSFER_SEMANTIC_COLUMNS
        if col in row.index
    )
    return any(keyword in text for keyword in ["调拨", "仓间", "调入"])


def _transfer_group_key(row, index):
    warehouse = _clean_transfer_text(row.get("仓库", "")).upper()
    trip_no = _clean_transfer_text(row.get("车次号", ""))
    batch_no = _clean_transfer_text(
        row.get("批次号", row.get("批次号集合", ""))
    )
    if trip_no:
        return f"TRIP||{warehouse}||{trip_no}"
    if batch_no:
        return f"BATCH||{warehouse}||{batch_no}"
    return f"ROW||{warehouse}||{index}"


def _ensure_transfer_columns(df):
    out = df.copy()
    text_columns = [
        "出库类型", "业务场景", "调入仓库", "实际目的地", "修正后目的地",
        "目的地", "标准地址", "标准邮编", "邮编前三位", "标准邮编集合",
        "邮编前三位集合", "目的州", "邮编来源", "FBA/FBX", "系统产品类型",
        "主产品类型", "平台名称", "FBA仓点代码", "FBX代码", "平台仓代码",
        "FBA仓点代码集合", "FBX代码集合", "平台仓代码集合", "平台仓配对集合",
        "批次目的地类型", "批次目的仓点", "目的仓点分配明细",
        "专线线路", "专线识别方式", "调拨目标仓代码", TRANSFER_AUDIT_COLUMN,
    ]
    for col in text_columns:
        if col not in out.columns:
            out[col] = ""
        else:
            out[col] = out[col].astype(object)
    for col in ["FBA出库体积", "FBX出库体积"]:
        if col not in out.columns:
            out[col] = 0.0
        else:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(float)
    if "目的地邮编待补充" not in out.columns:
        out["目的地邮编待补充"] = False
    else:
        out["目的地邮编待补充"] = out["目的地邮编待补充"].astype(object)
    return out


def _transfer_allocation_json(row, display):
    def numeric(name):
        value = pd.to_numeric(row.get(name, 0), errors="coerce")
        return 0.0 if pd.isna(value) else float(value)

    batch_no = _clean_transfer_text(
        row.get("批次号", row.get("批次号集合", ""))
    )
    return json.dumps(
        [{
            "对象类型": "其他",
            "平台": "盈仓",
            "仓点代码": display,
            "批次号": batch_no,
            "出库体积": numeric("出库体积"),
            "出库卡板数": numeric("出库卡板数"),
            "派送成本": numeric("派送成本"),
        }],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _clear_destination_for_invalid_transfer(out, indexes, audit_text):
    out.loc[indexes, "出库类型"] = "调拨"
    out.loc[indexes, "业务场景"] = "仓间调拨"
    out.loc[indexes, "调入仓库"] = ""
    out.loc[indexes, "调拨目标仓代码"] = ""
    out.loc[indexes, TRANSFER_AUDIT_COLUMN] = audit_text
    out.loc[indexes, "系统产品类型"] = "仓间调拨"
    out.loc[indexes, "主产品类型"] = "仓间调拨"
    out.loc[indexes, "FBA/FBX"] = ""
    out.loc[indexes, "平台名称"] = "盈仓"
    for col in TRANSFER_DESTINATION_TEXT_COLUMNS:
        out.loc[indexes, col] = ""
    for col in TRANSFER_RAW_CODE_COLUMNS + TRANSFER_BATCH_CODE_COLUMNS:
        out.loc[indexes, col] = ""
    out.loc[indexes, "FBA出库体积"] = 0.0
    out.loc[indexes, "FBX出库体积"] = 0.0
    out.loc[indexes, "批次目的地类型"] = ""
    out.loc[indexes, "批次目的仓点"] = ""
    out.loc[indexes, "目的仓点分配明细"] = ""
    out.loc[indexes, "标准邮编"] = ""
    out.loc[indexes, "邮编前三位"] = ""
    out.loc[indexes, "标准邮编集合"] = ""
    out.loc[indexes, "邮编前三位集合"] = ""
    out.loc[indexes, "目的州"] = ""
    out.loc[indexes, "邮编来源"] = ""
    out.loc[indexes, "目的地邮编待补充"] = True
    out.loc[indexes, "专线线路"] = ""
    out.loc[indexes, "专线识别方式"] = audit_text


def _apply_transfer_target(out, indexes, target, info, scope):
    display = info["display"]
    out.loc[indexes, "出库类型"] = "调拨"
    out.loc[indexes, "业务场景"] = "仓间调拨"
    out.loc[indexes, "调入仓库"] = display
    out.loc[indexes, "调拨目标仓代码"] = target
    out.loc[indexes, TRANSFER_AUDIT_COLUMN] = f"{scope}调拨统一覆盖:{display}"
    for col in TRANSFER_DESTINATION_TEXT_COLUMNS:
        out.loc[indexes, col] = display
    out.loc[indexes, "标准邮编"] = info["zip"]
    out.loc[indexes, "邮编前三位"] = info["zip3"]
    out.loc[indexes, "标准邮编集合"] = info["zip"]
    out.loc[indexes, "邮编前三位集合"] = info["zip3"]
    out.loc[indexes, "目的州"] = info["state"]
    out.loc[indexes, "邮编来源"] = "调拨目标仓地址规则"
    out.loc[indexes, "目的地邮编待补充"] = False
    out.loc[indexes, "FBA/FBX"] = ""
    out.loc[indexes, "系统产品类型"] = "仓间调拨"
    out.loc[indexes, "主产品类型"] = "仓间调拨"
    out.loc[indexes, "平台名称"] = "盈仓"
    for col in TRANSFER_RAW_CODE_COLUMNS + TRANSFER_BATCH_CODE_COLUMNS:
        out.loc[indexes, col] = ""
    out.loc[indexes, "平台仓代码集合"] = display
    out.loc[indexes, "FBA出库体积"] = 0.0
    out.loc[indexes, "FBX出库体积"] = 0.0
    out.loc[indexes, "批次目的地类型"] = "其他"
    out.loc[indexes, "批次目的仓点"] = display
    out.loc[indexes, "专线线路"] = info["line"]
    out.loc[indexes, "专线识别方式"] = "调拨目标仓优先覆盖"
    for index in indexes:
        out.at[index, "目的仓点分配明细"] = _transfer_allocation_json(
            out.loc[index],
            display,
        )


def apply_trip_transfer_destination_rules(df):
    """把调拨目的地作为清洗最高优先级，并按真实车次传播到全部批次。"""
    if df is None or df.empty:
        return df
    out = _ensure_transfer_columns(df)
    group_keys = pd.Series(
        (_transfer_group_key(row, index) for index, row in out.iterrows()),
        index=out.index,
        dtype=object,
    )
    for group_key, indexes in group_keys.groupby(group_keys, sort=False).groups.items():
        group = out.loc[indexes]
        transfer_mask = group.apply(transfer_row_has_semantics, axis=1)
        if not transfer_mask.any():
            continue

        targets = []
        for _, row in group.loc[transfer_mask].iterrows():
            for target in infer_transfer_targets_from_row(row):
                if target not in targets:
                    targets.append(target)
        if not targets:
            for _, row in group.iterrows():
                for target in infer_transfer_targets_from_row(row):
                    if target not in targets:
                        targets.append(target)

        if len(targets) == 0:
            _clear_destination_for_invalid_transfer(
                out,
                indexes,
                f"{TRANSFER_ERROR_PREFIX}调拨目标盈仓无法识别",
            )
            continue
        if len(targets) > 1:
            _clear_destination_for_invalid_transfer(
                out,
                indexes,
                f"{TRANSFER_ERROR_PREFIX}同车次调拨目标冲突:" + ",".join(targets),
            )
            continue

        scope = (
            "同车次"
            if str(group_key).startswith("TRIP||")
            else "同批次"
            if str(group_key).startswith("BATCH||")
            else "单行"
        )
        target = targets[0]
        _apply_transfer_target(
            out,
            indexes,
            target,
            TRANSFER_WAREHOUSE_INFO[target],
            scope,
        )
    return out


def transfer_override_error_rows(df):
    if df is None or df.empty or TRANSFER_AUDIT_COLUMN not in df.columns:
        return pd.DataFrame()
    audit = df[TRANSFER_AUDIT_COLUMN].fillna("").astype(str)
    return df[audit.str.startswith(TRANSFER_ERROR_PREFIX)].copy()


def apply_dominant_destination_from_detail(cleaned_batches, detail_df):
    """同一FTL车次混多个目的地时，用该车次内出库体积最大的明细行覆盖目的地识别字段。"""
    if cleaned_batches is None or cleaned_batches.empty or detail_df is None or detail_df.empty:
        return cleaned_batches
    detail = repair_numeric_columns(detail_df, columns=("出库体积", "出库卡板数", "派送成本"))
    if "批次号" not in detail.columns:
        return cleaned_batches
    out = ensure_object_df(cleaned_batches)
    destination_cols = [
        "系统产品类型", "主产品类型", "平台名称", "FBX代码集合", "平台仓代码集合", "平台仓配对集合",
        "FBA仓点代码集合", "标准邮编集合", "邮编前三位集合", "目的州", "邮编来源", "目的地邮编待补充",
    ]
    for col in destination_cols:
        if col not in out.columns:
            out[col] = ""

    for idx, row in out.iterrows():
        if str(row.get("主产品类型", "")) == "仓间调拨" or str(row.get("系统产品类型", "")) == "仓间调拨":
            continue
        batch_ids = split_values(row.get("批次号集合", row.get("批次号", "")))
        if not batch_ids:
            continue
        matched = detail[detail["批次号"].astype(str).isin(batch_ids)].copy()
        if matched.empty or "出库体积" not in matched.columns:
            continue
        matched["出库体积"] = pd.to_numeric(matched["出库体积"], errors="coerce").fillna(0)
        dominant = matched.sort_values("出库体积", ascending=False).iloc[0]

        product = str(dominant.get("FBA/FBX", dominant.get("系统产品类型", ""))).strip()
        if product not in ["FBA", "FBX"]:
            product = str(dominant.get("系统产品类型", row.get("主产品类型", ""))).strip()
        if product in ["FBA", "FBX"]:
            out.at[idx, "系统产品类型"] = product
            out.at[idx, "主产品类型"] = product

        zip_value = str(dominant.get("标准邮编", "")).strip()
        zip3_value = str(dominant.get("邮编前三位", "")).strip()
        state_value = str(dominant.get("目的州", "")).strip()
        if zip_value.lower() in ["nan", "none", "<na>"]:
            zip_value = ""
        if zip3_value.lower() in ["nan", "none", "<na>"]:
            zip3_value = ""
        if state_value.lower() in ["nan", "none", "<na>"]:
            state_value = ""

        out.at[idx, "标准邮编集合"] = zip_value
        out.at[idx, "邮编前三位集合"] = zip3_value or (zip_value[:3] if len(zip_value) == 5 else "")
        out.at[idx, "目的州"] = state_value
        out.at[idx, "邮编来源"] = str(dominant.get("邮编来源", "最大方数目的地覆盖")).strip() or "最大方数目的地覆盖"
        out.at[idx, "目的地邮编待补充"] = len(split_values(zip_value)) == 0

        fba_code = str(dominant.get("FBA仓点代码", "")).strip()
        platform = _normalize_platform_display(dominant.get("平台名称", ""))
        platform_code = _normalize_code_display(dominant.get("FBX代码", dominant.get("平台仓代码", dominant.get("仓库代码", ""))))
        if product == "FBA":
            out.at[idx, "FBA仓点代码集合"] = "" if fba_code.lower() in ["nan", "none", "<na>"] else _normalize_code_display(fba_code)
            out.at[idx, "平台名称"] = ""
            out.at[idx, "FBX代码集合"] = ""
            out.at[idx, "平台仓代码集合"] = ""
            out.at[idx, "平台仓配对集合"] = ""
        elif product == "FBX":
            out.at[idx, "FBA仓点代码集合"] = ""
            out.at[idx, "平台名称"] = platform
            out.at[idx, "FBX代码集合"] = platform_code
            out.at[idx, "平台仓代码集合"] = platform_code
            if platform and platform_code:
                out.at[idx, "平台仓配对集合"] = f"{platform}||{platform_code}"
    return out


def _normalize_platform_display(value):
    if processors.is_blank(value):
        return ""
    text = str(value).strip()
    if text.lower() in ["nan", "none", "null", "<na>"]:
        return ""
    key = re.sub(r"\s+", "", text).lower()
    return PLATFORM_DISPLAY_MAP.get(key, text)


def _normalize_code_display(value):
    if processors.is_blank(value):
        return ""
    text = str(value).strip()
    if text.lower() in ["nan", "none", "null", "<na>"]:
        return ""
    return text.upper()


def _normalize_code_list(value):
    values = split_values(value)
    out = []
    seen = set()
    for item in values:
        normalized = _normalize_code_display(item)
        key = normalized.upper()
        if normalized and key not in seen:
            out.append(normalized)
            seen.add(key)
    return ",".join(out)


def _normalize_pair_list(value):
    pairs = []
    seen = set()
    for raw in str(value or "").split(";"):
        if "||" not in raw:
            continue
        platform, code = raw.split("||", 1)
        platform = _normalize_platform_display(platform)
        code = _normalize_code_display(code)
        if not platform or not code:
            continue
        key = f"{platform.lower()}||{code.upper()}"
        if key not in seen:
            pairs.append(f"{platform}||{code}")
            seen.add(key)
    return ";".join(pairs)


def normalize_case_insensitive_labels(df):
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in PLATFORM_COLUMNS:
        if col in out.columns:
            out[col] = out[col].apply(_normalize_platform_display)
    for col in CODE_COLUMNS:
        if col in out.columns:
            out[col] = out[col].apply(_normalize_code_list if "集合" in col else _normalize_code_display)
    for col in PAIR_COLUMNS:
        if col in out.columns:
            out[col] = out[col].apply(_normalize_pair_list)
    return out


def _recalc_rank_and_share(df, value_col, rank_col, share_col, group_cols):
    if df.empty or value_col not in df.columns:
        return df
    out = df.copy()
    out[value_col] = pd.to_numeric(out[value_col], errors="coerce").fillna(0)
    out = out.sort_values(group_cols + [value_col], ascending=[True] * len(group_cols) + [False]).reset_index(drop=True)
    if share_col in out.columns:
        total = out.groupby(group_cols, dropna=False)[value_col].transform("sum")
        out[share_col] = out[value_col] / total.replace(0, pd.NA)
    if rank_col in out.columns:
        out[rank_col] = out.groupby(group_cols, dropna=False)[value_col].rank(method="first", ascending=False).astype(int)
    return out


def _group_sum_preserve_order(df, keys, sum_cols):
    if df.empty or not all(k in df.columns for k in keys):
        return df
    out = df.copy()
    for col in sum_cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    agg = {col: "sum" for col in sum_cols if col in out.columns}
    for col in out.columns:
        if col not in keys and col not in agg:
            agg[col] = "first"
    grouped = out.groupby(keys, dropna=False, as_index=False).agg(agg)
    return grouped[[c for c in out.columns if c in grouped.columns]]


def merge_case_duplicate_output_rows(df, sheet_type=""):
    if df is None or df.empty:
        return df
    out = normalize_case_insensitive_labels(df)

    if sheet_type == "FBX平台仓货量":
        keys = ["仓库", "统计周期", "平台", "平台仓代码"]
        out = _group_sum_preserve_order(out, keys, ["出库体积"])
        return _recalc_rank_and_share(out, "出库体积", "排名", "占比", ["仓库", "统计周期"])

    if sheet_type == "FBA货量排行":
        keys = ["仓库", "统计周期", "FBA仓点"]
        out = _group_sum_preserve_order(out, keys, ["出库体积"])
        return _recalc_rank_and_share(out, "出库体积", "排名", "占比", ["仓库", "统计周期"])

    if sheet_type == "成本":
        # 成本计算层已经基于业务筛选后的逐车明细生成平均值/P80。
        # 导出层只统一大小写和数值格式，不再聚合或用总量相除覆盖指标。
        return out

    return out


def clean_for_excel_output(df, sheet_type=""):
    if df is None:
        return pd.DataFrame()
    out = merge_case_duplicate_output_rows(df.copy(), sheet_type=sheet_type)
    for col in out.columns:
        if any(k in str(col) for k in TEXT_COLUMN_KEYWORDS):
            out[col] = out[col].fillna("").astype(str).replace({"nan": "", "None": "", "<NA>": ""})
    for col in INTEGER_OUTPUT_COLUMNS:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").round(0).astype("Int64")
    if "车次数" in out.columns:
        values = pd.to_numeric(out["车次数"], errors="coerce")
        if sheet_type in ["成本", "成本FTL", "分类型价格参考"]:
            out["车次数"] = values.round(2)
        else:
            out["车次数"] = values.round(0).astype("Int64")
    if sheet_type == "发车量" and "数值" in out.columns:
        out["数值"] = pd.to_numeric(out["数值"], errors="coerce").round(0).astype("Int64")
    for col in DECIMAL_OUTPUT_COLUMNS:
        if col in out.columns:
            if sheet_type == "发车量" and col == "数值":
                continue
            out[col] = pd.to_numeric(out[col], errors="coerce").round(2)
    return out


def _format_excel_ws(ws):
    if ws.max_row < 1:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        header = str(cell.value) if cell.value is not None else ""
        if any(k.upper() in header.upper() for k in TEXT_COLUMN_KEYWORDS):
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=cell.column).number_format = "@"
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)


def write_sheets_to_excel(sheets):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = str(sheet_name)[:31]
            cleaned = clean_for_excel_output(df, sheet_type=safe_name)
            cleaned.to_excel(writer, index=False, sheet_name=safe_name)
            _format_excel_ws(writer.book[safe_name])
    output.seek(0)
    return output


def safe_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r"[\\/:*?\"<>|\s]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "未命名"


def build_output_filename(warehouse, module_name, *descriptors, ext="xlsx"):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts = [warehouse, module_name]
    parts.extend([d for d in descriptors if d and str(d) not in ["请填入", "不适用", "全部"]])
    parts.append(timestamp)
    safe_parts = [safe_filename_part(p) for p in parts]
    return "_".join(safe_parts) + f".{ext}"
